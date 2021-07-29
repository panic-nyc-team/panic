# -*- coding: utf-8 -*-
from fuzzywuzzy import fuzz
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import traceback
import random
import openpyxl
import glob
from textblob import TextBlob
import text2emotion as te

from dateutil.relativedelta import relativedelta
from flask import Flask, render_template, request, url_for, flash, send_from_directory, send_file, jsonify
from werkzeug.utils import secure_filename, redirect
import json
import inspect, nltk
import numpy as np
import urllib.parse

from forms import (
    FileInputForm,
    PredictionDataForm,
    TrainModelForm,
    ChangeClassColorsForm,
    # SubmitAllForm,
    special_form
)

from util_functions import *

from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
import tensorflow.math as tfmath
import tensorflow.keras.backend as tfbackend

from os import mkdir
import sys
import requests
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_
import re
from flask_executor import Executor
import _thread
from timeloop import Timeloop
import datetime
from pytz import timezone
import urllib.parse

from sentence_transformers import SentenceTransformer, util

import webhoseio

import pandas as pd
from collections import Counter
from nltk.corpus import stopwords
import csv, pickle
from tqdm import tqdm
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize

ps = PorterStemmer()

from flask_cors import CORS

import io
import base64

from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import matplotlib

import spacy
from collections import Counter
import en_core_web_sm

nlp = en_core_web_sm.load()

webhoseio.config(token="8018e387-9258-4fd4-9ec5-9f9366a779a8")

app = Flask(__name__)
CORS(app)
app.config.from_object('config.Config')
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://mikenyc:12345@localhost/mike?charset=utf8'
# app.config ['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.sqlite3'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# COUNTRY_CODES = ['','US','AU','HK','GB']
# MARKET_LANGUAGE_CODES = ['','english','spanish','french']

COUNTRY_CODES = ["", "US", "AF", "AL", "DZ", "AS", "AD", "AO", "AI", "AQ", "AG", "AR", "AM", "AW", "AU", "AT", "AZ",
                 "BS", "BH", "BD", "BB", "BY", "BE", "BZ", "BJ", "BM", "BT", "BO", "BA", "BW", "BR", "IO", "VG", "BN",
                 "BG", "BF", "BI", "KH", "CM", "CA", "CV", "KY", "CF", "TD", "CL", "CN", "CX", "CC", "CO", "KM", "CK",
                 "CR", "HR", "CU", "CW", "CY", "CZ", "CD", "DK", "DJ", "DM", "DO", "TL", "EC", "EG", "SV", "GQ", "ER",
                 "EE", "ET", "FK", "FO", "FJ", "FI", "FR", "PF", "GA", "GM", "GE", "DE", "GH", "GI", "GR", "GL", "GD",
                 "GU", "GT", "GG", "GN", "GW", "GY", "HT", "HN", "HK", "HU", "IS", "IN", "ID", "IR", "IQ", "IE", "IM",
                 "IL", "IT", "CI", "JM", "JP", "JE", "JO", "KZ", "KE", "KI", "XK", "KW", "KG", "LA", "LV", "LB", "LS",
                 "LR", "LY", "LI", "LT", "LU", "MO", "MK", "MG", "MW", "MY", "MV", "ML", "MT", "MH", "MR", "MU", "YT",
                 "MX", "FM", "MD", "MC", "MN", "ME", "MS", "MA", "MZ", "MM", "NA", "NR", "NP", "NL", "AN", "NC", "NZ",
                 "NI", "NE", "NG", "NU", "KP", "MP", "NO", "OM", "PK", "PW", "PS", "PA", "PG", "PY", "PE", "PH", "PN",
                 "PL", "PT", "PR", "QA", "CG", "RE", "RO", "RU", "RW", "BL", "SH", "KN", "LC", "MF", "PM", "VC", "WS",
                 "SM", "ST", "SA", "SN", "RS", "SC", "SL", "SG", "SX", "SK", "SI", "SB", "SO", "ZA", "KR", "SS", "ES",
                 "LK", "SD", "SR", "SJ", "SZ", "SE", "CH", "SY", "TW", "TJ", "TZ", "TH", "TG", "TK", "TO", "TT", "TN",
                 "TR", "TM", "TC", "TV", "VI", "UG", "UA", "AE", "GB", "UY", "UZ", "VU", "VA", "VE", "VN", "WF", "EH",
                 "YE", "ZM", "ZW"]
MARKET_LANGUAGE_CODES = ["", "english", "afrikaans", "albanian", "amharic", "arabic", "armenian", "azerbaijani",
                         "basque", "belarusian", "bengali", "bulgarian", "burmese", "catalan", "cherokee", "chinese",
                         "chineset", "croatian", "czech", "danish", "dhivehi", "dutch", "estonian", "finnish", "french",
                         "galician", "georgian", "german", "greek", "gujarati", "hebrew", "hindi", "hungarian",
                         "icelandic", "ignore", "indonesian", "inuktitut", "irish", "italian", "japanese", "kannada",
                         "khmer", "korean", "laothian", "latvian", "lithuanian", "macedonian", "malay", "malayalam",
                         "maltese", "norwegian", "oriya", "persian", "polish", "portuguese", "punjabi", "romanian",
                         "russian", "serbian", "sinhalese", "slovak", "slovenian", "spanish", "swahili", "swedish",
                         "syriac", "tagalog", "tamil", "telugu", "thai", "tibetan", "turkish", "ukrainian", "urdu",
                         "vietnamese", "welsh", "yiddish"]

SITE_TYPES = ['', 'news', 'blogs', 'discussions']
REST = ['Day', 'Week', 'Month']

db = SQLAlchemy(app)
from models import *

db.create_all()
db.session.commit()
executor = Executor(app)

tl = Timeloop()
tz = timezone('EST')

GET = 'GET'
POST = 'POST'
TEST_STRING = ''

tokenizer = load_tokenizer()
model = load_model('./static/Models/model_under_use.h5')
maxlen = 40
class_colors = load_classColors()

sentence_model = SentenceTransformer('distilbert-base-nli-stsb-mean-tokens')


def startup():
    # reports = ReportModel.query.all()
    # for report in reports:
    #     # if os.path.exists(f'./static/jsons/report{report.id}.json'):
    #     #     continue
    #     print(report.id)
    #     try:
    #         d_f = report.date_from.strftime('%Y-%m-%d')
    #         d_t = report.date_to.strftime('%Y-%m-%d')
    #     except:
    #         d_f = None
    #         d_t = None
    #     if d_f is None and d_t is None:
    #         res = export_result(
    #             {'export_type': 'report', 'where': str(report.id), 'filter': 'includes',
    #              'search_parameter': '',
    #              'format': 'json', 'flag_link': True})
    #
    #     else:
    #         res = export_result(
    #             {'export_type': 'report', 'where': str(report.id), 'date_checkbox': 'date', 'filter': 'includes',
    #              'search_parameter': '', 'start_date': d_f, 'end_date': d_t,
    #              'format': 'json', 'flag_link': True})

    docs = NewDocumentModel.query.all()
    # counter = 1
    # for newdocument in docs:
    #     print(counter)
    #     counter += 1
        # if newdocument.domain_authority:
        #     continue
    offset = 0
    while True:
        l = []
        for c in docs[offset:offset+49]:
            if not c.domain_authority:
                l.append(c.url)
        if not l:
            offset += 49
            continue
        print(len(l))
        json_response = get_domain_authority(l)
        # print(json_response)
        results = json_response.get('results')
        if results:
            length = len(results)
            if length != len(l):
                print('not equal', length, len(docs))
                return
            i = 0
            while i < length:
                domain_authority = json_response['results'][i]['domain_authority']
                if domain_authority:
                    if domain_authority < 10:
                        bucket = '0-10'
                    elif domain_authority < 20:
                        bucket = '10-20'
                    elif domain_authority < 30:
                        bucket = '20-30'
                    elif domain_authority < 40:
                        bucket = '30-40'
                    elif domain_authority < 50:
                        bucket = '40-50'
                    elif domain_authority < 60:
                        bucket = '50-60'
                    elif domain_authority < 70:
                        bucket = '60-70'
                    elif domain_authority < 80:
                        bucket = '70-80'
                    elif domain_authority < 90:
                        bucket = '80-90'
                    elif domain_authority <= 100:
                        bucket = '90-100'
                    else:
                        domain_authority = -1
                        bucket = '-1'
                else:
                    domain_authority = -1
                    bucket = '-1'
                docs[offset+i].domain_authority = domain_authority
                docs[offset+i].bucket = bucket
                print(offset+i)
                i += 1
        db.session.commit()
        offset += 49
        time.sleep(10)
    # docs = NewDocumentModel.query.all()
    # print(len(docs))
    # for d in docs:
    #     try:
    #         if NewDocumentEntitiesModel.query.filter_by(f_id=d.id).first():
    #             continue
    #         article = nlp(d.text)
    #         items = [x.text for x in article.ents]
    #         a = Counter(items)
    #         for i in a:
    #             n = NewDocumentEntitiesModel(f_id=d.id, name=i, count=a[i])
    #             db.session.add(n)
    #         print('added', d.title)
    #         db.session.commit()
    #     except:
    #         print('1233123')
    #         db.session.rollback()

    search_queries = SuperSearchQueryModel.query.all()
    for search_query in search_queries:
        search_query.running = False
        docs = SearchQueryDocumentModel.query.filter_by(f_title=search_query.title).all()
        if docs:
            search_query.current_number = len(docs)
            search_query.total = len(docs)
        else:
            search_query.current_number = 0
            search_query.total = 0
    reports = ReportModel.query.all()
    for report in reports:
        report.running = False
        if not report.total:
            report.total = 0
        if not report.current_number:
            report.current_number = 0
    db.session.commit()


app.before_first_request(startup)


@app.after_request
def add_header(r):
    """
    Add headers to both force latest IE rendering engine or Chrome Frame,
    and also to cache the rendered page for 10 minutes.
    """
    r.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    r.headers["Pragma"] = "no-cache"
    r.headers["Expires"] = "0"
    r.headers['Cache-Control'] = 'public, max-age=0'
    return r


def get_sentiment(text):
    blob = TextBlob(text)
    sentiment = blob.sentiment.polarity
    if sentiment > .1:
        polarity = "Positive"
    elif sentiment < -.1:
        polarity = "Negative"
    else:
        polarity = "Neutral"
    return polarity, sentiment


# fl = 1
#
# @tl.job(interval=datetime.timedelta(seconds=10))
# def mints():
#     global fl
#     print(fl)
#     if fl == 0:
#         fl = 1
#         docs = SearchQueryDocumentModel.query.all()
#         c = 1
#         for d in docs:
#             if d.polarity:
#                 c += 1
#                 continue
#             if d.clean_text:
#                 polarity, sentiment = get_sentiment(d.clean_text)
#                 temp_emotions = te.get_emotion(d.clean_text)
#                 emotions = {}
#                 if temp_emotions:
#                     for key in temp_emotions:
#                         if temp_emotions[key] != 0.0:
#                             emotions[key] = temp_emotions[key]
#                 d.sentiment = sentiment
#                 d.polarity = polarity
#                 d.emotions = str(emotions)
#                 print(c)
#                 c += 1
#                 db.session.commit()
#         print('finish')
# sentences = SentenceTextModel.query.all()
# for s in sentences:
#     if s.polarity:
#         c += 1
#         continue
#     if s.sentence:
#         polarity, sentiment = get_sentiment(s.sentence)
#         temp_emotions = te.get_emotion(s.sentence)
#         emotions = {}
#         if temp_emotions:
#             for key in temp_emotions:
#                 if temp_emotions[key] != 0.0:
#                     emotions[key] = temp_emotions[key]
#         s.sentiment = sentiment
#         s.polarity = polarity
#         s.emotions = str(emotions)
#         # print(polarity)
#         # print(c)
#         c += 1
#         db.session.commit()

@tl.job(interval=datetime.timedelta(minutes=300))
def day():
    files = glob.glob('static/excel/*')
    for f in files:
        os.remove(f)

    work('Day')
    # report_word('Daily')


@tl.job(interval=datetime.timedelta(days=7))
def week():
    work('Week')
    # report_word('Weekly')


@tl.job(interval=datetime.timedelta(days=30))
def month():
    work('Month')
    # report_word('Monthly')


def work(fetch_frequency):
    searchqueries = SuperSearchQueryModel.query.filter_by(fetch_frequency=fetch_frequency, status='playing').all()
    if (searchqueries == []):
        print('Empty', file=sys.stderr)
        return None

    for searchquery in searchqueries:
        try:
            search_query_documents_background(searchquery.id)
        except Exception as e:
            print('error 123', e)


def report_word(frequency):
    reports = ReportModel.query.filter_by(frequency=frequency, up_to_date=True).all()
    if (len(reports) > 0):
        for i in reports:
            report_background(i.id, i.type, i.first, i.second, i.range_from, i.range_to)


@app.teardown_appcontext
def shutdown_session(exception=None):
    db.session.expunge_all()
    db.session.remove()


@app.route("/homeold", methods=[GET])
def homeold():
    return render_template('homeold.html')


@app.route("/", methods=[GET])
@app.route("/home", methods=[GET])
def home():
    return redirect(url_for("search_queries"))


@app.route("/utility", methods=[GET])
def utility():
    threshold = Threshold.query.filter_by(id=1).first()
    if (threshold is None):
        threshold = Threshold(id=1, value=100)
        db.session.add(threshold)
        db.session.commit()
    colors = ClassColors.query.filter_by(id=1).first()
    return render_template('utility.html', threshold=threshold.value, colors=colors)


@app.route("/setclasscolors", methods=[POST])
def setclasscolors():
    result = request.form
    narrative = result.get('narrative')
    if not narrative:
        narrative = ''
    aesthetic = result.get('aesthetic')
    if not aesthetic:
        aesthetic = ''
    craftsmanship = result.get('craftsmanship')
    if not craftsmanship:
        craftsmanship = ''
    purpose = result.get('purpose')
    if not purpose:
        purpose = ''
    overall = result.get('all')
    if not overall:
        overall = ''
    colors = ClassColors.query.filter_by(id=1).first()
    if colors is None:
        colors = ClassColors(id=1, narrative=narrative, aesthetic=aesthetic, craftsmanship=craftsmanship,
                             purpose=purpose, overall=overall)
        db.session.add(colors)
    else:
        colors.narrative = narrative
        colors.aesthetic = aesthetic
        colors.craftsmanship = craftsmanship
        colors.purpose = purpose
        colors.overall = overall
    db.session.commit()
    return redirect(url_for("utility"))


@app.route("/changeClassColors", methods=[GET, POST])
def change_class_colors():
    global class_colors

    form = ChangeClassColorsForm()
    if form.validate_on_submit():
        new_purpose = form.new_purpose.data
        new_craftsmaship = form.new_craftsmaship.data
        new_aesthetic = form.new_aesthetic.data
        new_narrative = form.new_narrative.data

        save_classColors(new_purpose, new_craftsmaship, new_aesthetic, new_narrative)
        class_colors = load_classColors()

        flash("Class Colors Changed Successfully!", "success")
        return redirect(url_for("utility"))

    return render_template("changeClassColors.html", form=form, class_colors=class_colors)


@app.route("/train", methods=[GET, POST])
def train():
    inputform = FileInputForm()

    if inputform.validate_on_submit():
        file = inputform.file.data
        if file.filename.split(".")[-1] != 'tsv':
            flash("ONLY UPLOAD A 'tsv' FILE!", "danger")
            return redirect(url_for('train'))

        singlefile(file)
        flash("File Successfully Uploaded", "success")
        file.close()
    trainModelform = TrainModelForm()

    return render_template("train.html", inputform=inputform, trainModelform=trainModelform)


@app.route('/train_model/<retrain>', methods=[GET, POST])
def train_model(retrain):
    global model, tokenizer, maxlen
    print(11112323123123)
    if retrain == 'True':
        # file_path = "bin/output2.tsv"
        data = loadTSVfromBin()
    else:
        # file_path = "static/File_Upload_Folder/uploaded.tsv"
        data = loadTSVfromFolder()
    try:
        ## data     = np.genfromtxt(file_path, delimiter='\t', dtype= str, encoding="utf8")
        # data = loadTSVfromFolder()
        data = np.array(data)

        features = data[:, 0]

        labels = data[:, 1]
        print("These are the model labels: ", labels)

        if len(features) < 70:
            flash("There must be atleast 70 rows of Data before training", "danger")
            return "less than 70"

        total_samples = data.shape[0]

        # Cleaning stop words and converting to lists
        features = filter_func(features)

        # shuffling the data
        features, labels = shuffle(features, labels)

        # Imp numbers to create Embeddings and for padding
        maxlen, count = count_words(features)
        num_words = len(count)
        maxlen = maxlen

        # One hot encoding Labels
        labels = onehot_encode_labels(labels)

        tok_features = tokenize(features, tokenizer)
        print("tok_features ------> ", tok_features)

        input_ids_in = tf.convert_to_tensor(tok_features[0])
        input_masks_in = tf.convert_to_tensor(tok_features[1])

        print("input_ids_in.shape ---> ", input_ids_in.shape)
        print("input_masks_in.shape --> ", input_masks_in.shape)

        # Getting Embeddings
        cls_token = embeddings(input_ids_in, input_masks_in)
        print("cls_token.shape ---> ", cls_token.shape)
        total_samples = len(cls_token)

        embedding_features = np.asarray(cls_token)

        # Training the Model
        model.fit(embedding_features, labels, epochs=50)

        # overwriting the model
        model.save('static/Models/model_under_use.h5')

        flash("Model Trained and Saved!", "success")
        return "training done"

    except ValueError as ve:
        flash("ERROR, Plz check if all your sentences end with a period i.e ' . '", "danger")
        print(ve)
        return "ValueError"

    except KeyError as ke:
        flash("ERROR, Encountered an Unknown Label during Training, Please Check Training Data", "danger")
        print(ke)
        return "keyError"

    except OSError as oe:
        flash("ERROR! No File uploded to Train on", "danger")
        print(oe)
        return "osError"

    except IndexError as ie:
        flash("There must be atleast 70 rows of Data before training", "danger")
        print(ie)
        return "indexError"
    except Exception as e:
        print(e)


@app.route("/restrat_model", methods=[POST])
def restart_model():
    global model

    model = load_model("static/Models/scratch_model.h5")
    model.save("static/Models/model_under_use.h5")

    flash("Model Started Form Scratch Successful", "success")

    return redirect(url_for('utility'))


@app.route("/test", methods=[GET, POST])
def test():
    global TEST_STRING

    predictionForm = PredictionDataForm()

    if predictionForm.validate_on_submit():
        TEST_STRING = predictionForm.text_area.data
        return redirect(url_for("results"))

    return render_template("test.html", predictionForm=predictionForm)


@app.route("/newtest", methods=[POST])
def newtest():
    global TEST_STRING
    text = request.form.get('text')
    if text:
        TEST_STRING = text
        return redirect(url_for("results"))
    else:
        return redirect(url_for('utility'))


@app.route("/results", methods=[GET, POST])
def results():
    global model, tokenizer, maxlen, TEST_STRING
    colors = ClassColors.query.filter_by(id=1).first()
    sentences = nltk.sent_tokenize(TEST_STRING)

    tok_test_features = tokenize(sentences, tokenizer)

    test_input_ids_in = tf.convert_to_tensor(tok_test_features[0])
    print("test_input_ids_in ----->", test_input_ids_in)
    test_input_masks_in = tf.convert_to_tensor(tok_test_features[1])
    print("test_input_ids_in --->", test_input_ids_in)

    # Getting Embeddings
    test_cls_token = embeddings(test_input_ids_in, test_input_masks_in)
    embedding_features = np.asarray(test_cls_token)

    predictions = model.predict(embedding_features)

    class_num = tfmath.argmax(predictions, axis=1)
    class_num = tfbackend.eval(class_num)
    labels = decode_onehot_labels(class_num)

    specialForm = special_form(labels)
    selects = [
        getattr(specialForm, f"special_{i}")
        for i in range(specialForm.n_attrs)
    ]

    data = list(zip(sentences, roundoff(predictions), labels, selects))
    bin_data = loadTSVfromBin()
    print("\n\nBIN LEN:", len(bin_data), '\n\n')

    if specialForm.validate_on_submit():
        corrected_labels = [
            sel.data
            for sel in selects
        ]

        appendTSVtoBin(corrected_labels, sentences)

        flash(
            f"Added {len(corrected_labels)} rows to the bin, Now total rows in bin are {len(bin_data) + len(corrected_labels)}",
            "success")
        return redirect(url_for("proceed"))

    return render_template(
        "results.html",
        data=data,
        len_data=len(data),
        bin_data=bin_data,
        len_bin_data=len(bin_data),
        colors=colors,
        specialForm=specialForm
    )


@app.route("/download_file", methods=[GET])
def download_file():
    path = "bin/output2.tsv"
    return send_file(path, as_attachment=True)


@app.route("/clear_bin", methods=[GET])
def clear_bin():
    clearBin()
    flash("Bin Emptied", "success")
    return redirect(url_for("utility"))


@app.route("/proceed", methods=[GET])
def proceed():
    return render_template("proceed.html")


@app.route("/predict", methods=["GET", "POST"])
def classified():
    text = request.json
    text = text['mytext']
    global model, tokenizer, maxlen

    sentences = nltk.sent_tokenize(text)
    tok_test_features = tokenize(sentences, tokenizer)

    test_input_ids_in = tf.convert_to_tensor(tok_test_features[0])
    # print("test_input_ids_in ----->", test_input_ids_in)
    test_input_masks_in = tf.convert_to_tensor(tok_test_features[1])
    # print("test_input_ids_in --->", test_input_ids_in)
    # Getting Embeddings
    test_cls_token = embeddings(test_input_ids_in, test_input_masks_in)
    embedding_features = np.asarray(test_cls_token)

    predictions = model.predict(embedding_features)

    class_num = tfmath.argmax(predictions, axis=1)  # Returns the index with the largest value across axes of a tensor.
    class_num = tfbackend.eval(class_num)
    labels = decode_onehot_labels(class_num)

    dict = {

    }
    dat = zip(sentences, labels)
    for i in dat:
        dict[i[0]] = i[1]
    return json.dumps(dict)


@app.route("/getsimilarity", methods=['POST'])
def sentenceSimliarity():
    dimensions = {'aesthetic': 0, 'craftsmanship': 0, 'purpose': 0, 'narrative': 0}
    data = request.get_json(force=True)
    sentence1 = data.get('sentence1')
    sentence2 = data.get('sentence2')
    a = getSimlarity(sentence1, sentence2)
    return jsonify(a)


def getSimlarity(sentence1, sentence2):
    s1 = []
    s2 = []
    for i in sentence1:
        s1.append(i)
    for i in sentence2:
        s2.append(i[0])

    embeddings1 = sentence_model.encode(s1, convert_to_tensor=True)
    embeddings2 = sentence_model.encode(s2, convert_to_tensor=True)
    # Compute cosine-similarits
    cosine_scores = util.pytorch_cos_sim(embeddings1, embeddings2)

    # Output the pairs with their score
    d = {}
    for i in range(len(sentence1)):
        # s = []
        dd = {}
        for j in range(len(sentence2)):
            # print(sentence1[i],sentence2[j])
            # s.append("{:.2f}".format(cosine_scores[i][j]))
            dd[sentence2[j][0]] = {'similarity': "{:.2f}".format(cosine_scores[i][j]), 'id': sentence2[j][1],
                                   'type': sentence2[j][2], 'title': sentence2[j][3], 'url': sentence2[j][4]}
        d[sentence1[i]] = dd
    return d


@app.route("/addToBin", methods=["GET", "POST"])
def addToBin():
    text = request.json
    lab = text['labels']
    sen = text['sentences']

    x = zip(sen, lab)
    json_data_tuples = list(x)

    bin_data = loadTSVfromBin()

    for i in json_data_tuples:
        for j in bin_data:
            # comparing only sentences
            if i[0] == j[0]:
                bin_data.remove(j)

    bin_data_labels = []
    bin_data_sentences = []
    for i in bin_data:
        bin_data_sentences.append(i[0])
        bin_data_labels.append(i[1])

    writeTSVtoBin(bin_data_labels, bin_data_sentences)
    appendTSVtoBin(lab, sen)
    return ("bin_data")


@app.route("/seeBin", methods=[GET, POST])
def seeBin():
    bin_data = loadTSVfromBin()
    print(bin_data)
    bin_labels = []
    bin_sentences = []
    for i in bin_data:
        bin_labels.append(i[1])
        bin_sentences.append(i[0])

    specialForm = special_form(bin_labels)
    selects = [
        getattr(specialForm, f"special_{i}")
        for i in range(specialForm.n_attrs)
    ]

    if specialForm.validate_on_submit():
        corrected_labels = [
            sel.data
            for sel in selects
        ]

        writeTSVtoBin(corrected_labels, bin_sentences)

    data = list(zip(bin_sentences, bin_labels, selects))
    threshold = Threshold.query.filter_by(id=1).first()
    if (threshold is None):
        threshold = Threshold(id=1, value=100)
        db.session.add(threshold)
        db.session.commit()
    return render_template(
        "seeBin.html",
        data=data,
        bin_data=bin_data,
        len_bin_data=len(bin_data),
        class_colors=class_colors,
        specialForm=specialForm,
        threshold=threshold.value
    )


@app.route('/export')
def export():
    reports = ReportModel.query.all()
    search_queries = SuperSearchQueryModel.query.all()
    attributes = ['title', 'author', 'publish_date', 'site', 'site_type', 'url', 'main_image',
                  'country', 'text']
    return render_template('export.html', reports=reports, search_queries=search_queries, attributes=attributes)


@app.route('/exportresult', methods=['post'])
def export_result(temp_form=None):
    if temp_form:
        form = temp_form
    else:
        form = request.form
    print(form)
    flag_link = form.get('flag_link')
    if not flag_link:
        flag_link = False
    else:
        flag_link = True
    export_type = form.get('export_type')
    id = form.get('where')
    format = form.get('format')
    start_date = form.get('start_date')
    end_date = form.get('end_date')
    date_checkbox = form.get('date_checkbox')
    if date_checkbox == 'date':
        date_checkbox = True
    else:
        date_checkbox = None
    field_checkbox = form.get('field_checkbox')
    if field_checkbox == 'field':
        field_checkbox = True
    else:
        field_checkbox = None
    if date_checkbox:
        try:
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        except Exception as e:
            print(e)
            return 'date error'
    if export_type == 'document':
        id = 'document'
    if id is None:
        return 'id error'
    if export_type == 'report':
        report = ReportModel.query.filter_by(id=id).first()
        if (report):
            data = []
            sentences = SentenceModel.query.filter_by(f_id=report.id).all()
            sentence_text = {}
            for i in SentenceTextModel.query.filter_by(f_id=report.id).all():
                if not i.sentence:
                    continue
                try:
                    # polarity, sentiment = get_sentiment(i.sentence)
                    # temp_emotions = te.get_emotion(i.sentence)
                    # emotions = {}
                    # if temp_emotions:
                    #     for key in temp_emotions:
                    #         if temp_emotions[key] != 0.0:
                    #             emotions[key] = temp_emotions[key]
                    sentence_text[i.id] = {'text': i.sentence, 'polarity': i.polarity, 'emotions': i.emotions}
                except Exception as e:
                    print(e)
            if (report.type == 'vscompany'):
                c1 = CompanyDocumentModel.query.filter_by(title=report.first).first()
                c2 = CompanyDocumentModel.query.filter_by(title=report.second).first()
                for s in sentences:
                    if c1 and c2:
                        try:
                            s1 = sentence_text.get(int(s.sentence1))
                            s2 = sentence_text.get(int(s.sentence2))
                            data.append({'s1': {'text': s1.get('text'),
                                                'polarity': s1.get('polarity'),
                                                'emotions': s1.get('emotions'),
                                                'parent_title': c1.title,
                                                'parent_url': url_for('edit_company', title=c1.title, _external=True)},
                                         's2': {'text': s2.get('text'), 'polarity': s2.get('polarity'),
                                                'emotions': s2.get('emotions'), 'parent_title': c2.title,
                                                'parent_url': url_for('edit_company', title=c2.title, _external=True)},
                                         'similarity_dimension': s.dimension, 'similarity': s.similarity})
                        except:
                            pass
            elif report.type == 'vssearchquery':
                c1 = CompanyDocumentModel.query.filter_by(title=report.first).first()
                for s in sentences:
                    c2 = NewDocumentModel.query.filter_by(f_id=s.id2).first()
                    if date_checkbox:
                        if c1 and c2:
                            try:
                                published = datetime.datetime.strptime(c2.published.split('T')[0], '%Y-%m-%d').date()
                            except:
                                print('date split error')
                                continue
                            if published and start_date <= published <= end_date:
                                try:
                                    s1 = sentence_text.get(int(s.sentence1))
                                    s2 = sentence_text.get(int(s.sentence2))
                                    data.append(
                                        {'s1': {'text': s1.get('text'), 'polarity': s1.get('polarity'),
                                                'emotions': s1.get('emotions'), 'parent_title': c1.title
                                            , 'parent_url': url_for('edit_company', title=c1.title, _external=True)}
                                            , 's2': {'text': s2.get('text'), 'polarity': s2.get('polarity'),
                                                     'emotions': s2.get('emotions'), 'url': c2.url,
                                                     'parent_title': c2.title
                                            , 'parent_url': url_for('edit_search_query_document', id=s.id2,
                                                                    _external=True)
                                            , 'parent_date': c2.published, 'parent_site': c2.site
                                            , 'parent_site_type': c2.site_type, 'parent_country': c2.country}
                                            , 'similarity_dimension': s.dimension, 'similarity': s.similarity})
                                except:
                                    pass
                            else:
                                continue
                    else:
                        if c1 and c2:
                            try:
                                s1 = sentence_text.get(int(s.sentence1))
                                s2 = sentence_text.get(int(s.sentence2))
                                data.append(
                                    {'s1': {'text': s1.get('text'), 'polarity': s1.get('polarity'),
                                            'emotions': s1.get('emotions'), 'parent_title': c1.title
                                        , 'parent_url': url_for('edit_company', title=c1.title, _external=True)}
                                        , 's2': {'text': s2.get('text'), 'polarity': s2.get('polarity'),
                                                 'emotions': s2.get('emotions'), 'url': c2.url,
                                                 'parent_title': c2.title
                                        , 'parent_url': url_for('edit_search_query_document', id=s.id2, _external=True)
                                        , 'parent_date': c2.published, 'parent_site': c2.site
                                        , 'parent_site_type': c2.site_type, 'parent_country': c2.country}
                                        , 'similarity_dimension': s.dimension, 'similarity': s.similarity})
                            except:
                                pass

            elif report.type == 'vstag':
                c1 = CompanyDocumentModel.query.filter_by(title=report.first).first()
                for s in sentences:
                    c2 = CompanyDocumentModel.query.filter_by(title=s.title2).first()
                    if (c2):
                        if c1 and c2:
                            try:
                                s1 = sentence_text.get(int(s.sentence1))
                                s2 = sentence_text.get(int(s.sentence2))
                                data.append(
                                    {'s1': {'text': s1.get('text'), 'polarity': s1.get('polarity'),
                                            'emotions': s1.get('emotions'), 'parent_title': c1.title
                                        , 'parent_url': url_for('edit_company', title=c1.title, _external=True)}
                                        ,
                                     's2': {'text': s2.get('text'), 'polarity': s2.get('polarity'),
                                            'emotions': s2.get('emotions'), 'parent_title': c2.title
                                         , 'parent_url': url_for('edit_company', title=c2.title, _external=True)}
                                        , 'similarity_dimension': s.dimension,
                                     'similarity': s.similarity})
                            except:
                                pass

                    else:
                        c2 = ArbitraryDocumentModel.query.filter_by(title=s.title2).first()
                        if c1 and c2:
                            try:
                                s1 = sentence_text.get(int(s.sentence1))
                                s2 = sentence_text.get(int(s.sentence2))
                                data.append(
                                    {'s1': {'text': s1.get('text'), 'polarity': s1.get('polarity'),
                                            'emotions': s1.get('emotions'), 'parent_title': c1.title
                                        , 'parent_url': url_for('edit_company', title=c1.title, _external=True)}
                                        ,
                                     's2': {'text': s2.get('text'), 'polarity': s2.get('polarity'),
                                            'emotions': s2.get('emotions'), 'parent_title': c2.title
                                         , 'parent_url': url_for('edit_arbitrary_document', title=c2.title,
                                                                 _external=True),
                                            'url': c2.url
                                         , 'author': c2.author, 'provider': c2.provider, 'image_url': c2.image_url}
                                        , 'similarity_dimension': s.dimension,
                                     'similarity': s.similarity})
                            except:
                                pass

            else:
                return 'error'
            if format == 'flat_json':
                result_flat = []
                for d in data:
                    flat_attrs_temp = {}
                    temps1 = d.get('s1')
                    if temps1:
                        for i in temps1:
                            if i == 'emotions':
                                try:
                                    temp_em = eval(temps1[i])
                                except:
                                    continue
                                if temp_em:
                                    for j in temp_em:
                                        flat_attrs_temp[f's1_{i}_{j}'] = temp_em.get(j)
                            else:
                                flat_attrs_temp['s1_' + i] = temps1[i]

                    temps2 = d.get('s2')
                    if temps2:
                        for i in temps2:
                            if i == 'emotions':
                                try:
                                    temp_em = eval(temps2[i])
                                except:
                                    continue
                                if temp_em:
                                    for j in temp_em:
                                        flat_attrs_temp[f's2_{i}_{j}'] = temp_em.get(j)
                            else:
                                flat_attrs_temp['s2_' + i] = temps2[i]
                    flat_attrs_temp['similarity_dimension'] = d.get('similarity_dimension')
                    flat_attrs_temp['similarity'] = d.get('similarity')
                    result_flat.append(flat_attrs_temp.copy())
                return jsonify(result_flat)
            if format == 'excel':
                flat_attrs = {}
                wb = openpyxl.Workbook()
                sheet = wb.active
                if not data:
                    return 'empty'
                column_map = {}
                temps1 = data[0].get('s1')
                if temps1:
                    for i in temps1:
                        flat_attrs['s1_' + i] = True
                temps2 = data[0].get('s2')
                if temps2:
                    for i in temps2:
                        flat_attrs['s2_' + i] = True
                flat_attrs['similarity_dimension'] = True
                flat_attrs['similarity'] = True
                for count, i in enumerate(flat_attrs, 1):
                    row_1 = sheet.cell(row=1, column=count)
                    row_1.value = i
                    column_map[i] = count
                result_flat = []
                for d in data:
                    flat_attrs_temp = {}
                    temps1 = d.get('s1')
                    if temps1:
                        for i in temps1:
                            flat_attrs_temp['s1_' + i] = temps1[i]
                    temps2 = d.get('s2')
                    if temps2:
                        for i in temps2:
                            flat_attrs_temp['s2_' + i] = temps2[i]
                    flat_attrs_temp['similarity_dimension'] = d.get('similarity_dimension')
                    flat_attrs_temp['similarity'] = d.get('similarity')
                    result_flat.append(flat_attrs_temp.copy())
                for row, i in enumerate(result_flat, 2):
                    for j in i:
                        if column_map.get(j):
                            row_1 = sheet.cell(row=row, column=column_map.get(j))
                            row_1.value = i.get(j)
                name = f'static/excel/result{random.randint(0, 999)}.xlsx'
                while os.path.exists(name):
                    name = f'static/excel/result{random.randint(0, 999)}.xlsx'
                wb.save(name)
                return send_file(name, as_attachment=True)

            if flag_link:
                name = f'./static/jsons/report{report.id}.json'
                # if os.path.exists(name):
                #     os.remove(name)
                print(data)
                with open(name, 'w') as f:
                    json.dump(data, f)
                return {'report_id': report.id}
            return jsonify(data)
        else:
            return 'report not found'
    elif export_type == 'search_query':
        search_query = SuperSearchQueryModel.query.filter_by(id=id).first()
        documents = NewDocumentModel.query.filter_by(f_title=search_query.title).all()
        data = get_doc_data(documents, field_checkbox, form, date_checkbox, start_date, end_date)
        if format == 'excel':
            if data == 'empty':
                return 'empty'
            return send_file(data, as_attachment=True)
        if flag_link:
            name = f'./static/jsons/searchquery{search_query.id}.json'
            with open(name, 'w') as f:
                json.dump(data, f)
            return {'report_id': search_query.id}
        return jsonify(data)

    elif export_type == 'document':
        documents = NewDocumentModel.query.all()
        data = get_doc_data(documents, field_checkbox, form, date_checkbox, start_date, end_date)
        if format == 'excel':
            if data == 'empty':
                return 'empty'
            return send_file(data, as_attachment=True)
        return jsonify(data)
    else:
        return 'error'
    # return 'error'


def get_doc_data(documents, field_checkbox, form, date_checkbox, start_date, end_date):
    data = []
    search_p = None
    attributes = {}
    format = form.get('format')
    if field_checkbox:
        attributes = {'title': None, 'author': None, 'publish_date': None, 'site': None
            , 'site_type': None, 'url': None, 'main_image': None, 'country': None, 'text': None}
        for attr in attributes:
            if form.get(attr):
                attributes[attr] = attr
        search_p = form.get('search_parameter')
    for document in documents:
        if date_checkbox:
            try:
                published = datetime.datetime.strptime(document.published.split('T')[0], '%Y-%m-%d').date()
            except:
                print('date split error')
                continue
            if published and start_date <= published <= end_date:
                temp = get_doc_sub(document, field_checkbox, search_p, attributes, format)
                if temp == 'return':
                    continue
                data.append(temp)
            else:
                continue
        else:
            temp = get_doc_sub(document, field_checkbox, search_p, attributes, format)
            if temp == 'return':
                continue
            data.append(temp)
    if format == 'flat_json':
        result_flat = []
        for i in data:
            for j in i:
                result_flat.append(j)
        return result_flat
    elif format == 'excel':
        result_flat = []
        for i in data:
            for j in i:
                # print(j)
                result_flat.append(j)
        largest = 0
        largest_index = None
        for index, i in enumerate(result_flat):
            if len(i) > largest:
                largest = len(i)
                largest_index = index
        if largest == 0:
            return 'empty'
        wb = openpyxl.Workbook()
        sheet = wb.active
        column_map = {}
        for count, i in enumerate(result_flat[largest_index], 1):
            # print(i)
            row_1 = sheet.cell(row=1, column=count)
            row_1.value = i
            column_map[i] = count
        for row, i in enumerate(result_flat, 2):
            for j in i:
                if column_map.get(j):
                    row_1 = sheet.cell(row=row, column=column_map.get(j))
                    row_1.value = i[j]

        # c2 = sheet.cell(row=1, column=2)
        # c2.value = "RAI"
        name = f'static/excel/result{random.randint(0, 999)}.xlsx'
        while os.path.exists(name):
            name = f'static/excel/result{random.randint(0, 999)}.xlsx'
        wb.save(name)
        return name

    elif format == 'json':
        return data
    else:
        return 'error'


def get_doc_sub(document, field_checkbox, search_p, attributes, format):
    # print(1)
    # polarity, sentiment = get_sentiment(document.text)
    # temp_emotions = te.get_emotion(document.text)
    # emotions = {}
    # if temp_emotions:
    #     for key in temp_emotions:
    #         if temp_emotions[key] != 0.0:
    #             emotions[key] = temp_emotions[key]
    d = SearchQueryDocumentModel.query.filter_by(id=document.f_id).first()
    temp = {'search_query': document.f_title, 'title': document.title, 'author': document.author
        , 'publish_date': document.published, 'site': document.site,
            'site_type': document.site_type
        , 'url': document.url, 'main_image': document.main_image,
            'country': document.country
        , 'text': document.text}
    emotions = None
    if d:
        if d.polarity:
            temp['polarity'] = d.polarity
        else:
            temp['polarity'] = ''
        if d.emotions:
            emotions = eval(d.emotions)
    # print(2)
    if field_checkbox:
        if search_p:
            if search_p not in document.title and search_p not in document.text:
                return 'continue'
        for attr in attributes:
            if attributes[attr] is None:
                temp.pop(attr, None)
    # print(3)
    # d_persons = NewDocumentPersonsModel.query.filter_by(f_id=document.id).all()
    # d_locations = NewDocumentLocationsModel.query.filter_by(f_id=document.id).all()
    # d_organizations = NewDocumentOrganizationsModel.query.filter_by(f_id=document.id).all()
    persons = []
    locations = []
    organizations = []
    if format == 'flat_json' or format == 'excel':
        result = []
        # if not d_persons and not d_locations and not d_organizations:
        #     result.append(temp.copy())
        if not emotions:
            result.append(temp.copy())
        else:
            for i in emotions:
                t = temp.copy()
                t['emotion'] = i
                result.append(t.copy())
            # for i in d_organizations:
            #     if emotions:
            #         for j in emotions:
            #             t = temp.copy()
            #             t['entity'] = i.name
            #             t['sentiment'] = i.sentiment
            #             t['emotion'] = j
            #             result.append(t.copy())
            #     else:
            #         t = temp.copy()
            #         t['entity'] = i.name
            #         t['sentiment'] = i.sentiment
            #         result.append(t.copy())
        # print(4)
        return result
    elif format == 'json':
        # for i in d_persons:
        #     persons.append({'name': i.name, 'sentiment': i.sentiment})
        # for i in d_locations:
        #     locations.append({'name': i.name, 'sentiment': i.sentiment})
        # for i in d_organizations:
        #     organizations.append({'name': i.name, 'sentiment': i.sentiment})
        # temp['entities'] = {'persons': persons, 'organizations': organizations, 'locations': locations}
        if emotions:
            temp['emotions'] = emotions
        else:
            temp['emotions'] = {}
        # temp['reach'] = {'per_million': document.reach_per_m, 'page_views': {'per_million': document.reach_views_per_m
        #     , 'per_user': document.reach_views_per_u}}
        print(4)
        return temp
    else:
        return 'error'


@app.route('/highlight', methods=['POST'])
def highlight():
    # id = request.form.get('id')
    url = request.form.get('url')
    sentence = request.form.get('sentence')
    print(url, sentence)
    # search_query_document = SearchQueryDocumentModel.query.filter_by(id=id).first()
    if (url):
        url = url + '#:~:text=' + urllib.parse.quote(sentence)
        return url
    else:
        return 'error'


@app.route('/seewebhose')
def see_webhose():
    id = request.args.get('id')
    if (not id):
        return 'id error'
    data = get_webhose(id)
    if data == 'error':
        return 'no document found'
    return render_template('seewebhose.html', data=data.replace('\n', '<br>'))


def get_webhose(id):
    newdocument = NewDocumentModel.query.filter_by(id=id).first()
    if (newdocument is None):
        return 'error'
    d_categories = NewDocumentSiteCategoriesModel.query.filter_by(f_id=id).all()
    d_links = NewDocumentExternalLinksModel.query.filter_by(f_id=id).all()
    d_images = NewDocumentExternalImagesModel.query.filter_by(f_id=id).all()
    # d_persons = NewDocumentPersonsModel.query.filter_by(f_id=id).all()
    # d_locations = NewDocumentLocationsModel.query.filter_by(f_id=id).all()
    # d_organizations = NewDocumentOrganizationsModel.query.filter_by(f_id=id).all()
    site_categories = []
    external_links = []
    external_images = []
    persons = []
    locations = []
    organizations = []
    for i in d_categories:
        site_categories.append(i.category)
    for i in d_links:
        external_links.append(i.url)
    for i in d_images:
        external_images.append(i.url)
    # for i in d_persons:
    #     persons.append('{{ name: "{name}",  sentiment: "{sentiment}" }}'.format(name=i.name, sentiment=i.sentiment))
    # for i in d_locations:
    #     locations.append('{{ name: "{name}",  sentiment: "{sentiment}" }}'.format(name=i.name, sentiment=i.sentiment))
    # for i in d_organizations:
    #     organizations.append(
    #         '{{ name: "{name}",  sentiment: "{sentiment}" }}'.format(name=i.name, sentiment=i.sentiment))

    data = '''
    {{ \n
        thread: {{

            uuid: "{thread_uuid}" ,

            url: "{url}" ,

            site_full: "{site_full}" ,

            site: "{site}" ,

            site_section: "{site_section}" ,

            site_categories: {site_categories} ,

            section_title: "{section_title}" ,

            title: "{title}" ,

            title_full: "{title_full}" ,

            published: "{published}" ,

            replies_count: {replies_count} ,

            participants_count: {participants_count} ,

            site_type: "{site_type}" ,

            country: "{country}" ,

            spam_score: {spam_score} ,

            main_image: "{main_image}" ,

            performance_score: {performance_score} ,

            domain_rank: {domain_rank} ,

            reach: {{

            per_million: {reach_per_m} ,

            page_views: {{

            per_million: {reach_views_per_m} ,

            per_user: {reach_views_per_u}

            }} ,

            updated: "{reach_updated}"

            }} ,

            social: {{

            facebook: {{

            likes: {facebook_likes} ,

            comments: {facebook_comments} ,

            shares: {facebook_shares}

            }} ,

            gplus: {{

            shares: {gplus_shares}

            }} ,

            pinterest: {{

            shares: {pinterest_shares}

            }} ,

            linkedin: {{

            shares: {linkedin_shares}

            }} ,

            stumbledupon: {{

            shares: {stumbledupon_shares}

            }} ,

            vk: {{

            shares: {vk_shares}

            }}

            }}

            }} ,

        uuid: "{uuid}" ,

        url: "{url}" ,

        ord_in_thread: {ord_in_thread} ,

        parent_url: {parent_url} ,

        author: "{author}" ,

        published: "{published}" ,

        title: "{title}" ,

        text: "{text}" ,

        highlightText: "{highlight_text}" ,

        highlightTitle: "{highlight_title}" ,

        highlightThreadTitle: "{highlight_thread_title}" ,

        language: "{language}" ,

        external_links: {external_links} ,

        external_images: {external_images},

        rating: {rating} ,

        crawled: "{crawled}" ,

        updated: "{updated}"

    }} 
        '''.format(thread_uuid=newdocument.thread_uuid, url=newdocument.url, site_full=newdocument.site_full
                   , site=newdocument.site, site_section=newdocument.site_section, site_categories=str(site_categories)
                   , section_title=newdocument.section_title, title=newdocument.title, title_full=newdocument.title_full
                   , published=newdocument.published, replies_count=newdocument.replies_count
                   , participants_count=newdocument.participants_count, site_type=newdocument.site_type
                   , country=newdocument.country, spam_score=newdocument.spam_score, main_image=newdocument.main_image
                   , performance_score=newdocument.performance_score, domain_rank=newdocument.domain_rank
                   , reach_per_m=newdocument.reach_per_m, reach_views_per_m=newdocument.reach_views_per_m
                   , reach_views_per_u=newdocument.reach_views_per_u, reach_updated=newdocument.reach_updated
                   , facebook_likes=newdocument.facebook_likes, facebook_comments=newdocument.facebook_comments
                   , facebook_shares=newdocument.facebook_shares, gplus_shares=newdocument.gplus_shares
                   , pinterest_shares=newdocument.pinterest_shares, linkedin_shares=newdocument.linkedin_shares
                   , stumbledupon_shares=newdocument.stumbledupon_shares, vk_shares=newdocument.vk_shares,
                   uuid=newdocument.uuid
                   , ord_in_thread=newdocument.ord_in_thread, parent_url=newdocument.parent_url,
                   author=newdocument.author
                   , text=newdocument.text, highlight_text=newdocument.highlight_text,
                   highlight_title=newdocument.highlight_title
                   , highlight_thread_title=newdocument.highlight_thread_title, language=newdocument.language
                   , external_links=str(external_links), external_images=str(external_images)
                   , rating=newdocument.rating
                   , crawled=newdocument.crawled, updated=newdocument.updated)
    return data


@app.route('/threshold')
def threshold():
    threshold = Threshold.query.filter_by(id=1).first()
    if (threshold):
        return render_template('threshold.html', threshold=threshold.value)
    else:
        return render_template('threshold.html', threshold=100)


@app.route('/setthreshold')
def set_threshold():
    threshold = Threshold.query.filter_by(id=1).first()
    t = request.args.get('t')
    if (t is None or t == ''):
        return 'error'
    if (threshold is None):
        threshold = Threshold(id=1, value=t)
        db.session.add(threshold)
        db.session.commit()
    else:
        threshold.value = t
        db.session.commit()
    SentenceModel.query.filter(SentenceModel.similarity < threshold.value).delete()
    reports = ReportModel.query.all()
    for report in reports:
        update_score(report.id)
    return redirect(url_for('utility'))


@app.route('/<type>/savetobin', methods=['POST'])
def save_to_bin(type):
    dimensions = ["aesthetic", "narrative", "craftsmanship", "purpose"]
    title = request.form.get('title')
    id = request.form.get('id')
    sentence = request.form.get('sentence')
    if (title is None or title == '' or title == 'None') and (id is None or id == ''):
        return 'error'
    temp = None
    if (type == 'companies'):
        temp = CompanyDocumentModel.query.filter_by(title=title).first()
    elif (type == 'searchquerydocuments'):
        temp = SearchQueryDocumentModel.query.filter_by(id=id).first()
        if (temp is None):
            temp = SearchQueryDocumentModel.query.filter(SearchQueryDocumentModel.f_title == title).filter(
                SearchQueryDocumentModel.classified_sentences.contains(sentence)).first()
    elif (type == 'arbitrarydocuments'):
        temp = ArbitraryDocumentModel.query.filter_by(title=title).first()
    elif (type == 'tags'):
        comp_temp = CompanyDocumentModel.query.filter_by(title=title).first()
        if (comp_temp):
            temp = comp_temp
        else:
            arb_temp = ArbitraryDocumentModel.query.filter_by(title=title).first()
            temp = arb_temp
    else:
        return 'error 0'
    if (temp and temp.classified_sentences):
        sentences = eval(temp.classified_sentences)
        url = url_for('addToBin', _external=True)
        headers = {'Content-type': 'application/json', 'Accept': 'text/plain'}
        labels_data = []
        sentences_data = []
        for i in sentences:
            if (sentences[i] in dimensions):
                labels_data.append(sentences[i])
                sentences_data.append(i)
        if (len(labels_data) == len(sentences_data)):
            myobj = {'labels': labels_data, 'sentences': sentences_data}
            myobj = json.dumps(myobj)
            x = requests.post(url, data=myobj, headers=headers)
            if (x.ok):
                return 'done'
            else:
                return 'error 1'
        else:
            return 'error 2'
    else:
        return 'error 3'


@app.route('/<type>/saveclassifier', methods=['POST'])
def save_classifier(type):
    dimensions = ["aesthetic", "narrative", "craftsmanship", "purpose"]
    title = request.form.get('title')
    id = request.form.get('id')
    dimension = request.form.get('dimension')
    sentence = request.form.get('sentence')
    if (
            title is None or title == '' or sentence is None or sentence == '' or dimension is None or dimension not in dimensions):
        return 'error'
    temp = None
    if (type == 'companies'):
        temp = CompanyDocumentModel.query.filter_by(title=title).first()
    elif (type == 'searchquerydocuments'):
        temp = SearchQueryDocumentModel.query.filter_by(id=id).first()
        if (temp is None):
            temp = SearchQueryDocumentModel.query.filter(SearchQueryDocumentModel.f_title == title).filter(
                SearchQueryDocumentModel.classified_sentences.contains(sentence)).first()
    elif (type == 'arbitrarydocuments'):
        temp = ArbitraryDocumentModel.query.filter_by(title=title).first()
    elif (type == 'tags'):
        comp_temp = CompanyDocumentModel.query.filter_by(title=title).first()
        if (comp_temp):
            temp = comp_temp
        else:
            arb_temp = ArbitraryDocumentModel.query.filter_by(title=title).first()
            temp = arb_temp
    else:
        return 'error 1'
    if (temp):
        if (temp and temp.classified_sentences):
            sentences = eval(temp.classified_sentences)
            if (sentences.get(sentence)):
                sentences[sentence] = dimension
                temp.classified_sentences = str(sentences)
        else:
            return 'cannot find company'
        db.session.commit()
    else:
        return 'error 11'
    return 'done'


@app.route('/<type>/classifier', methods=['POST'])
def classifier(type):
    sentences = None
    temp_colors = ClassColors.query.filter_by(id=1).first()
    if temp_colors:
        colors = {'purpose': temp_colors.purpose, 'craftsmanship': temp_colors.craftsmanship,
                  'aesthetic': temp_colors.aesthetic, 'narrative': temp_colors.narrative}
    else:
        colors = {'purpose': '', 'craftsmanship': '',
                  'aesthetic': '', 'narrative': ''}
    title = request.form.get('title')
    id = request.form.get('query_document_id')
    highlight_sentence = request.form.get('sentence')
    if (id is None):
        if (title is None or title == ''):
            return 'query_document_id is empty'
    else:
        if (id is None or id == ''):
            return 'query_document_id is empty'
    if (type == 'companies'):
        temp = CompanyDocumentModel.query.filter_by(title=title).first()
        if (temp and temp.classified_sentences):
            sentences = eval(temp.classified_sentences)
            clean_text = temp.clean_text
        else:
            return 'cannot find company'
    elif (type == 'searchquerydocuments'):
        print(id, title, highlight_sentence)
        if (id):
            temp = SearchQueryDocumentModel.query.filter_by(id=id).first()
        else:
            temp = SearchQueryDocumentModel.query.filter(SearchQueryDocumentModel.f_title == title).filter(
                SearchQueryDocumentModel.classified_sentences.contains(highlight_sentence)).first()
        if (temp and temp.classified_sentences):
            sentences = eval(temp.classified_sentences)
            clean_text = temp.clean_text
        else:
            return 'error'
    elif (type == 'arbitrarydocuments'):
        temp = ArbitraryDocumentModel.query.filter_by(title=title).first()
        if (temp and temp.classified_sentences):
            sentences = eval(temp.classified_sentences)
            clean_text = temp.clean_text
        else:
            return 'error'
    elif (type == 'tags'):
        temp = None
        comp_temp = CompanyDocumentModel.query.filter_by(title=title).first()
        if comp_temp:
            temp = comp_temp
        else:
            arb_temp = ArbitraryDocumentModel.query.filter_by(title=title).first()
            if arb_temp:
                temp = arb_temp

        # for i in CompanyDocumentModel.query.all():
        #     if(i.industry_tags and title in i.industry_tags and i.classified_sentences and highlight_sentence in i.classified_sentences):
        #         temp = i
        #         break
        # if(temp is None):
        #     for i in ArbitraryDocumentModel.query.all():
        #         if(i.industry_tags and title in i.industry_tags and i.classified_sentences and highlight_sentence in i.classified_sentences):
        #             temp = i
        #             break
        if (temp and temp.classified_sentences):
            sentences = eval(temp.classified_sentences)
            clean_text = temp.clean_text
        else:
            return 'error'
    else:
        return 'error 1'
    dimensions = ["aesthetic", "narrative", "craftsmanship", "purpose"]
    return render_template('classifier.html', sentences=sentences, dimensions=dimensions, title=title,
                           highlight_sentence=highlight_sentence, class_colors=colors, id=id,
                           clean_text=clean_text)


@app.route('/industrytags/')
def industry_tags_route():
    try:
        industry_tags = IndustryTags.query.all()
        return render_template('industrytags.html', industry_tags=industry_tags)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/industrytags/add', methods=['POST'])
def industry_tags_add():
    try:
        result = request.form
        company_industry_tags = result.get('company_industry_tags')
        arbitrary_industry_tags = result.get('arbitrary_industry_tags')
        if (company_industry_tags):
            old_title = result.get('old_title')
            if (old_title is None or old_title == ''):
                return 'error'
            companydocument = CompanyDocumentModel.query.filter_by(title=old_title).first()
            if (companydocument):
                if (companydocument.industry_tags is None or companydocument.industry_tags == ''):
                    companydocument.industry_tags = company_industry_tags
                else:
                    if (company_industry_tags in companydocument.industry_tags):
                        return 'industry tag already exists in this company'
                    companydocument.industry_tags += ',' + company_industry_tags
                db.session.commit()
                return redirect(url_for('edit_company') + '?title=' + old_title)
        elif (arbitrary_industry_tags):
            old_title = result.get('old_title')
            if (old_title is None or old_title == ''):
                return 'error'
            arbitrarydocument = ArbitraryDocumentModel.query.filter_by(title=old_title).first()
            if (arbitrarydocument):
                if (arbitrarydocument.industry_tags is None or arbitrarydocument.industry_tags == ''):
                    arbitrarydocument.industry_tags = arbitrary_industry_tags
                else:
                    if (arbitrary_industry_tags in arbitrarydocument.industry_tags):
                        return 'industry tag already exists in this company'
                    arbitrarydocument.industry_tags += ',' + arbitrary_industry_tags
                db.session.commit()
                return redirect(url_for('edit_arbitrary_document') + '?title=' + old_title)

        else:
            title = result.get('title')
            if (title and title != ''):
                industry_tag = IndustryTags(title=title)
            else:
                return 'empty title'
            db.session.add(industry_tag)
            db.session.commit()
            return redirect(url_for('industry_tags_route'))
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/industrytags/edit', methods=['POST', 'GET'])
def industry_tags_edit():
    if (request.method == 'POST'):
        # try:
        result = request.form
        print(result, file=sys.stderr)
        title = result.get('title')
        old_title = result.get('old_title')

        if (old_title and old_title != ''):
            industry_tag = IndustryTags.query.filter_by(title=old_title).first()
        else:
            return 'empty title 0'
        if (industry_tag):
            if (not title or title == ''):
                return 'empty title 1'
            industry_tag.title = title
            db.session.commit()
            companydocuments = CompanyDocumentModel.query.all()
            print(companydocuments, type(companydocuments), file=sys.stderr)
            arbitrarydocuments = ArbitraryDocumentModel.query.all()
            if (len(companydocuments) > 0):
                for companydocument in companydocuments:
                    if (companydocument.industry_tags and old_title in companydocument.industry_tags):
                        companydocument.industry_tags = companydocument.industry_tags.replace(old_title, title)
                    db.session.commit()
            if (len(arbitrarydocuments) > 0):
                for arbitrarydocument in arbitrarydocuments:
                    if (arbitrarydocument.industry_tags and old_title in arbitrarydocument.industry_tags):
                        arbitrarydocument.industry_tags = arbitrarydocument.industry_tags.replace(old_title, title)
                    db.session.commit()
        else:
            return 'empty title 2'
        return redirect(url_for('industry_tags_route'))
    # except Exception as e:
    #     print(e,file=sys.stderr)
    #     return 'error'
    else:
        try:
            title = request.args.get('title')
            if (title and title != ''):
                industry_tag = IndustryTags.query.filter_by(title=title).first()
                if (not industry_tag):
                    return 'error'
            else:
                return 'error'
            return render_template('editindustrytag.html', industry_tag=industry_tag)
        except Exception as e:
            print(e, file=sys.stderr)
            return 'error'


@app.route('/industrytags/delete', methods=['POST'])
def industry_tags_delete():
    try:
        result = request.form
        company_industry_tags = result.get('company_industry_tags')
        arbitrary_industry_tags = result.get('arbitrary_industry_tags')
        if (company_industry_tags):
            old_title = result.get('old_title')
            title = result.get('title')
            if (old_title is None or old_title == ''):
                return 'error 1'
            companydocument = CompanyDocumentModel.query.filter_by(title=old_title).first()
            if (companydocument):
                list_tags = companydocument.industry_tags.split(',')
                print(company_industry_tags, file=sys.stderr)
                list_tags.remove(company_industry_tags)
                list_tags = ','.join(list_tags)
                if (list_tags == ''):
                    list_tags = None
                companydocument.industry_tags = list_tags
                db.session.commit()
                return redirect(url_for('edit_company') + '?title=' + old_title)
        elif (arbitrary_industry_tags):
            old_title = result.get('old_title')
            title = result.get('title')
            if (old_title is None or old_title == ''):
                return 'error 2'
            arbitrarydocument = ArbitraryDocumentModel.query.filter_by(title=old_title).first()
            if (arbitrarydocument):
                list_tags = arbitrarydocument.industry_tags.split(',')
                print(arbitrary_industry_tags, file=sys.stderr)
                list_tags.remove(arbitrary_industry_tags)
                list_tags = ','.join(list_tags)
                if (list_tags == ''):
                    list_tags = None
                arbitrarydocument.industry_tags = list_tags
                db.session.commit()
            return redirect(url_for('edit_arbitrary_document') + '?title=' + old_title)
        else:
            title = result.get('title')
            if (title == '' or title is None):
                return 'no search query document found'
            if (IndustryTags.delete(title=title)):
                companydocuments = CompanyDocumentModel.query.all()
                arbitrarydocuments = ArbitraryDocumentModel.query.all()
                if (len(companydocuments) > 0):
                    for companydocument in companydocuments:
                        if (companydocument.industry_tags and title in companydocument.industry_tags):
                            list_tags = companydocument.industry_tags.split(',')
                            list_tags.remove(title)
                            list_tags = ','.join(list_tags)
                            if (list_tags == ''):
                                list_tags = None
                            companydocument.industry_tags = list_tags
                            db.session.commit()
                if (len(arbitrarydocuments) > 0):
                    for arbitrarydocument in arbitrarydocuments:
                        if (arbitrarydocument.industry_tags and title in arbitrarydocument.industry_tags):
                            list_tags = arbitrarydocument.industry_tags.split(',')
                            list_tags.remove(title)
                            list_tags = ','.join(list_tags)
                            if (list_tags == ''):
                                list_tags = None
                            arbitrarydocument.industry_tags = list_tags
                            db.session.commit()
                    return redirect(url_for('industry_tags_route'))
            else:
                return 'error 3'
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error 4'


@app.route('/searchqueries/deletesearchquerydocument')
def delete_search_query_document():
    id = request.args.get('id')
    if (id == '' or id is None):
        return 'no search query document found'
    try:
        f_title = SearchQueryDocumentModel.query.filter_by(id=id).first().f_title
        if (SearchQueryDocumentModel.delete(id=id)):
            doc = NewDocumentModel.query.filter_by(f_id=id).first()
            NewDocumentModel.delete(id=doc.id)
            NewDocumentEntitiesModel.delete(f_id=doc.id)
            # NewDocumentOrganizationsModel.delete(f_id=doc.id)
            # NewDocumentPersonsModel.delete(f_id=doc.id)
            # NewDocumentLocationsModel.delete(f_id=doc.id)
            NewDocumentSiteCategoriesModel.delete(f_id=doc.id)
            NewDocumentExternalLinksModel.delete(f_id=doc.id)
            NewDocumentExternalImagesModel.delete(f_id=doc.id)
            return redirect(url_for('search_query_documents', title=f_title))
        else:
            return 'error'
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/searchqueries/seeclassifiedsentences', methods=['POST'])
def search_queries_classified_sentences():
    result = request.form
    title = result.get('title')
    if (title is None or title == ''):
        return "title cannot be empty"
    try:
        searchquerydocument = SearchQueryDocumentModel.query.filter_by(title=title).first()
        if (searchquerydocument is None):
            return 'Error'
        d = eval(searchquerydocument.classified_sentences)
        return render_template('seeclassifiedsentences.html',
                               classified_sentences=searchquerydocument.classified_sentences)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/searchqueries/savesearchquerydocument', methods=['POST'])
def save_search_query_document():
    try:
        result = request.form
        clean_text = result.get('clean_text')
        searchquerydocument = SearchQueryDocumentModel(title=result.get('title'), clean_text=clean_text,
                                                       date=result.get('date'), author=result.get('author'),
                                                       provider=result.get('provider'), url=result.get('url'),
                                                       image_url=result.get('image_url'))
        old_title = result.get('old_title')
        id = result.get('id')
        if (id is None or id == ''):
            return "Error"
        if (searchquerydocument.clean_text == '' or searchquerydocument.clean_text == 'None'):
            searchquerydocument.clean_text = None
        if (searchquerydocument.clean_text is not None):
            # searchquerydocument.classified_sentences = str(temp_azure(searchquerydocument.clean_text))
            # res = requests.post(url_for('classified',_external=True), json={"mytext":re.sub(r'[^a-zA-Z0-9. ]','',searchquerydocument.clean_text)})
            res = requests.post(url_for('classified', _external=True), json={"mytext": searchquerydocument.clean_text})
            if res.ok:
                searchquerydocument.classified_sentences = str(res.json())
            else:
                searchquerydocument.classified_sentences = None
        else:
            searchquerydocument.classified_sentences = None
        polarity, sentiment = get_sentiment(clean_text)
        temp_emotions = te.get_emotion(clean_text)
        emotions = {}
        if temp_emotions:
            for key in temp_emotions:
                if temp_emotions[key] != 0.0:
                    emotions[key] = temp_emotions[key]
        edit_document = SearchQueryDocumentModel.query.filter_by(id=id).first()
        edit_document.title = searchquerydocument.title
        edit_document.clean_text = searchquerydocument.clean_text
        edit_document.classified_sentences = str(searchquerydocument.classified_sentences)
        edit_document.date = searchquerydocument.date
        edit_document.author = searchquerydocument.author
        edit_document.provider = searchquerydocument.provider
        edit_document.url = searchquerydocument.url
        edit_document.image_url = searchquerydocument.image_url
        edit_document.sentiment = sentiment
        edit_document.polarity = polarity
        edit_document.emotions = str(emotions)
        # SearchQueryDocumentModel.query.filter_by(title = old_title).update(dict(title=searchquerydocument.title,clean_text=searchquerydocument.clean_text,classified_sentences=str(searchquerydocument.classified_sentences),date=searchquerydocument.date,author=searchquerydocument.author,provider=searchquerydocument.provider,url=searchquerydocument.url,image_url=searchquerydocument.image_url))
        db.session.commit()
        return redirect(url_for('search_query_documents', title=SearchQueryDocumentModel.query.filter_by(
            title=searchquerydocument.title).first().f_title))
    except Exception as e:
        print(e, file=sys.stderr)
        return "Error"


@app.route('/searchqueries/changestatus', methods=['POST'])
def search_query_change_status():
    title = request.form.get('title')
    print(title, file=sys.stderr)
    if (title is None or title == ''):
        return 'error'
    searchquery = SuperSearchQueryModel.query.filter_by(title=title).first()
    print(title, file=sys.stderr)
    if (searchquery.status == 'playing'):
        searchquery.status = 'paused'
    elif (searchquery.status == 'paused'):
        searchquery.status = 'playing'
    else:
        return 'error'
    try:
        db.session.commit()
    except:
        return 'error'
    return 'done'


@app.route('/searchqueries/editsearchquerydocument')
def edit_search_query_document():
    id = request.args.get('id')
    if (id == '' or id is None):
        return 'no search query document found'
    searchquerydocument = SearchQueryDocumentModel.query.filter_by(id=id).first()
    newdocument = NewDocumentModel.query.filter_by(f_id=id).first()
    if (searchquerydocument is not None):
        return render_template('editsearchquerydocument.html', searchquerydocument=searchquerydocument,
                               newdocument=newdocument)
    else:
        return 'Error'


@app.route('/searchqueries/searchquerydocuments')
def search_query_documents():
    data = ''
    title = request.args.get('title')
    if (title == '' or title is None):
        return 'no search query document found'
    sort = request.args.get('sort')
    searchquerydocuments = []
    for i in SearchQueryDocumentModel.query.filter_by(f_title=title).all():
        searchquerydocuments.append(i)
    if (sort == 'titleup'):
        searchquerydocuments.sort(key=lambda x: x.title.lower(), reverse=True)
    elif (sort == 'titledown'):
        searchquerydocuments.sort(key=lambda x: x.title.lower(), reverse=False)
    elif (sort == 'sourceup'):
        searchquerydocuments.sort(key=lambda x: x.provider.lower(), reverse=True)
    elif (sort == 'sourcedown'):
        searchquerydocuments.sort(key=lambda x: x.provider.lower(), reverse=False)
    elif (sort == 'dateup'):
        searchquerydocuments.sort(key=lambda x: x.date, reverse=True)
    elif (sort == 'datedown'):
        searchquerydocuments.sort(key=lambda x: x.date, reverse=False)
    elif (sort == 'processingup'):
        temp = []
        for i in searchquerydocuments:
            if (i.clean_text and len(i.clean_text) > 500):
                temp.append([i, True])
            else:
                temp.append([i, False])
        temp.sort(key=lambda x: x[1], reverse=True)
        searchquerydocuments = []
        for i in temp:
            searchquerydocuments.append(i[0])
    elif (sort == 'processingdown'):
        temp = []
        for i in searchquerydocuments:
            if (i.clean_text and len(i.clean_text) > 500):
                temp.append([i, True])
            else:
                temp.append([i, False])
        temp.sort(key=lambda x: x[1], reverse=False)
        searchquerydocuments = []
        for i in temp:
            searchquerydocuments.append(i[0])

    return render_template('searchquerydocuments.html', searchquerydocuments=searchquerydocuments, title=title)


@app.route('/search', methods=['POST'])
def search():
    search = request.form.get('q')
    title = request.form.get('title')
    # search = 'example first lowercase article'
    search = search.lower()
    search = search.split(' ')
    query = []
    search = [k for k in search if k != '']
    clauses = [SearchQueryDocumentModel.title.like('%{0}%'.format(k)) for k in search]
    temp_articles = SearchQueryDocumentModel.query.filter(or_(*clauses)).all()
    articles = []
    for i in temp_articles:
        t = []
        heading = i.title.lower().split(' ')
        heading = [k for k in heading if k != '']
        flag = False
        for s in heading:
            for q in search:
                if (s == q):
                    flag = True
                    # print(s,q,file=sys.stderr)
        if (flag):
            t.append(i)
            t.append(0)
            for s in heading:
                for q in search:
                    if (s == q):
                        t[1] += 1
            articles.append(t)
    articles.sort(key=lambda x: x[1], reverse=True)
    result = []
    for i in articles:
        if (i[0].f_title == title):
            result.append(i[0])
    return render_template('searchquerydocuments.html', searchquerydocuments=result, title=title)


@app.route('/searchqueries/')
def search_queries():
    try:
        searchqueries = SuperSearchQueryModel.query.all()
        dates = []
        for s in searchqueries:
            temp = SearchQueryModel.query.filter_by(f_id=s.id).first()
            if temp:
                dates.append(temp)
            else:
                dates.append(None)
        total = []
        yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
        new_documents = []
        for i in searchqueries:
            total.append(len(SearchQueryDocumentModel.query.filter_by(f_title=i.title).all()))
            new_documents.append(len(
                SearchQueryDocumentModel.query.filter(SearchQueryDocumentModel.f_title == i.title).filter(
                    SearchQueryDocumentModel.date_created > yesterday).all()))
        return render_template('searchqueries.html', searchqueries=searchqueries, total=total,
                               new_documents=new_documents,
                               dates=dates, zip=zip)
    except Exception as e:
        print(e, 1, file=sys.stderr)
        return 'error'


@app.route('/searchqueries/addnewsearchquery')
def add_new_search_query():
    # return render_template('addeditsearchqueries.html', COUNTRY_CODES=COUNTRY_CODES,
    #                        MARKET_LANGUAGE_CODES=MARKET_LANGUAGE_CODES, SITE_TYPES=SITE_TYPES, REST=REST,
    #                        searchquery=SearchQueryModel(title='', query_string='', characters='', site=''))
    return render_template('supersearchqueries.html', COUNTRY_CODES=COUNTRY_CODES,
                           MARKET_LANGUAGE_CODES=MARKET_LANGUAGE_CODES, SITE_TYPES=SITE_TYPES, REST=REST,
                           supersearchquery=SuperSearchQueryModel(title=''), searchqueries=[])


@app.route('/searchqueries/editsearchquery')
def edit_search_queries():
    title = request.args.get('title')
    if (title == '' or title is None):
        return 'no search query found'
    try:
        supersearchquery = SuperSearchQueryModel.query.filter_by(title=title).first()
        searchqueries = SearchQueryModel.query.filter_by(f_id=supersearchquery.id).all()
        return render_template('supersearchqueries.html', COUNTRY_CODES=COUNTRY_CODES,
                               MARKET_LANGUAGE_CODES=MARKET_LANGUAGE_CODES, SITE_TYPES=SITE_TYPES, REST=REST,
                               searchqueries=searchqueries, supersearchquery=supersearchquery)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/searchqueries/deletesearchquery')
def delete_search_query():
    title = request.args.get('title')
    if (title == '' or title is None):
        return 'no search query found'
    try:
        sq = SuperSearchQueryModel.query.filter_by(title=title).first()
        if SearchQueryModel.delete(f_id=sq.id) and SearchQueryDocumentModel.deleteall(f_title=title):
            SuperSearchQueryModel.query.filter_by(title=title).delete()
            doc_list = NewDocumentModel.query.filter_by(f_title=title).all()
            NewDocumentModel.deleteall(f_title=title)
            for doc in doc_list:
                NewDocumentEntitiesModel.delete(f_id=doc.id)
                # NewDocumentOrganizationsModel.delete(f_id=doc.id)
                # NewDocumentPersonsModel.delete(f_id=doc.id)
                # NewDocumentLocationsModel.delete(f_id=doc.id)
                NewDocumentSiteCategoriesModel.delete(f_id=doc.id)
                NewDocumentExternalLinksModel.delete(f_id=doc.id)
                NewDocumentExternalImagesModel.delete(f_id=doc.id)
            db.session.commit()
            return redirect(url_for('search_queries'))
        else:
            return 'error'
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/searchqueries/savesearchquery', methods=['POST'])
def savesearchquery():
    try:
        result = request.form
        old_title = result.get('old_title')
        title = result.get('title')
        fetch_frequency = result.get('fetch_frequency')
        atts = []
        for i in range(1, 30):
            query_string = result.get(f'query_string{i}')
            market_language_code = result.get(f'market_language_code{i}')
            country_code = result.get(f'country_code{i}')
            # site_type = result.get(f'site_type{i}')
            site = result.get(f'site{i}')
            characters = result.get(f'characters{i}')
            if query_string:
                atts.append({'query_string': query_string, 'market_language_code': market_language_code,
                             'country_code': country_code, 'site': site, 'characters': characters})
        freshness = 30
        if (title is None or title == ''):
            return "title cannot be empty"
        # if (query_string is None or query_string == ''):
        #     return "query string cannot be empty"
        # if (site == '' or site == None):
        #     site = ''
        # if (characters == '' or characters == None):
        #     characters = ''

        super_search_query = SuperSearchQueryModel.query.filter_by(title=old_title).first()
        if super_search_query:
            super_search_query.title = title
            super_search_query.running = True
        else:
            t = SuperSearchQueryModel(title=title, fetch_frequency=fetch_frequency, status='playing', total=0,
                                      current_number=0, running=True)
            db.session.add(t)
        db.session.commit()

        super_search_query = SuperSearchQueryModel.query.filter_by(title=title).first()
        SearchQueryModel.query.filter_by(f_id=super_search_query.id).delete()
        query_documents = SearchQueryDocumentModel.query.filter_by(f_title=old_title).all()
        for d in query_documents:
            d.f_title = title
        new_documents = NewDocumentModel.query.filter_by(f_title=old_title).all()
        for d in new_documents:
            d.f_title = title
        reports_temp = ReportModel.query.filter_by(second=old_title).all()
        for d in reports_temp:
            d.second = title
        for i in atts:
            searchquery = SearchQueryModel(query_string=i.get('query_string'),
                                           market_language_code=i.get('market_language_code'),
                                           country_code=i.get('country_code'),
                                           site=i.get('site'), freshness=freshness, characters=i.get('characters'),
                                           f_id=super_search_query.id)
            db.session.add(searchquery)
            db.session.commit()
        executor.submit(search_query_documents_background, super_search_query.id)
        db.session.expunge_all()
        db.session.close()

        return redirect(url_for('search_queries'))

    except Exception as e:
        print(e, file=sys.stderr)
        return "Error"


@app.route('/arbitrarydocuments/')
def arbitrary_documents():
    try:
        documents = ArbitraryDocumentModel.query.all()
    except Exception as e:
        print(e, file=sys.stderr)
        documents = []
    return render_template('arbitrarydocuments.html', documents=documents)


@app.route('/arbitrarydocuments/addnew')
def add_new_document():
    try:
        arbitrarydocument = ArbitraryDocumentModel(title='', author='', provider='', url='', image_url='', date='',
                                                   industry_tags=None, clean_text='', classified_sentences='')
        industrytags = IndustryTags.query.all()
        return render_template('addeditarbitrarydocument.html', arbitrarydocument=arbitrarydocument,
                               industrytags=industrytags)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/arbitrarydocuments/edit')
def edit_arbitrary_document():
    result = request.args
    title = result.get('title')
    title = urllib.parse.unquote(title)
    print(title, file=sys.stderr)
    if (title == '' or title is None):
        return 'no arbitrary document found'
    try:
        arbitrarydocument = ArbitraryDocumentModel.query.filter_by(title=title).first()
        industrytags = IndustryTags.query.all()
        return render_template('addeditarbitrarydocument.html', arbitrarydocument=arbitrarydocument,
                               industrytags=industrytags)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/arbitrarydocuments/savearbitrarydocument', methods=['POST'])
def save_arbitrary_document():
    try:
        result = request.form
        clean_text = result.get('clean_text')
        # if(clean_text):
        #     clean_text = re.sub(r"(\r\n|\r|\n)","",clean_text)
        #     clean_text = re.sub(r"([^0-9]\.)",r"\1 ",clean_text)
        arbitrarydocument = ArbitraryDocumentModel(title=result.get('title'), clean_text=clean_text,
                                                   date=result.get('date'), author=result.get('author'),
                                                   provider=result.get('provider'), url=result.get('url'),
                                                   image_url=result.get('image_url'),
                                                   industry_tags=result.get('industry_tags'))
        old_title = result.get('old_title')
        if (arbitrarydocument.title is None or arbitrarydocument.title == ''):
            return "title cannot be empty"
        if (arbitrarydocument.clean_text == '' or arbitrarydocument.clean_text == 'None'):
            arbitrarydocument.clean_text = None
        if (arbitrarydocument.clean_text is not None):
            # arbitrarydocument.classified_sentences = str(temp_azure(arbitrarydocument.clean_text))
            # res = requests.post(url_for('classified',_external=True), json={"mytext":re.sub(r'[^a-zA-Z0-9. ]','',arbitrarydocument.clean_text)})
            res = requests.post(url_for('classified', _external=True), json={"mytext": arbitrarydocument.clean_text})
            if res.ok:
                arbitrarydocument.classified_sentences = str(res.json())
            else:
                arbitrarydocument.classified_sentences = None
        else:
            arbitrarydocument.classified_sentences = None
        if (old_title is not None and old_title != ''):
            ArbitraryDocumentModel.query.filter_by(title=old_title).update(
                dict(title=arbitrarydocument.title, clean_text=arbitrarydocument.clean_text,
                     classified_sentences=str(arbitrarydocument.classified_sentences), date=arbitrarydocument.date,
                     author=arbitrarydocument.author, provider=arbitrarydocument.provider, url=arbitrarydocument.url,
                     image_url=arbitrarydocument.image_url, industry_tags=arbitrarydocument.industry_tags),
            )
            db.session.commit()
        else:
            db.session.add(arbitrarydocument)
            db.session.commit()
        return redirect(url_for('arbitrary_documents'))
    except Exception as e:
        db.session.rollback()
        print(e, file=sys.stderr)
        return "Error"


@app.route('/arbitrarydocuments/delete')
def delete_arbitrary_document():
    title = request.args.get('title')
    if (title == '' or title is None):
        return 'no arbitrary document found'
    try:
        if (ArbitraryDocumentModel.delete(title=title)):
            return redirect(url_for('arbitrary_documents'))
        else:
            return 'delete error'
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/arbitrarydocuments/seeclassifiedsentences')
def see_classified_sentences_arbitrary():
    title = request.args.get('title')
    try:
        arbitrarydocument = ArbitraryDocumentModel.query.filter_by(title=title).first()
        d = eval(arbitrarydocument.classified_sentences)
        return render_template('seeclassifiedsentences.html',
                               classified_sentences=arbitrarydocument.classified_sentences)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/companies/')
def companies():
    try:
        tag = request.args.get('tag')
        companies = CompanyDocumentModel.query.all()
        industry_tags = IndustryTags.query.all()
        if (tag and tag != 'all'):
            for i in companies[:]:
                if (i.industry_tags is None or tag not in i.industry_tags):
                    companies.remove(i)
        return render_template('companies.html', companies=companies, industry_tags=industry_tags,
                               eval=eval)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/companies/addnewcompanies')
def add_new_companies():
    try:
        searchqueries = SuperSearchQueryModel.query.all()
        industrytags = IndustryTags.query.all()
        return render_template('addeditcompanies.html', companydocument=CompanyDocumentModel(title='', clean_text=''),
                               searchqueries=searchqueries, report=None, industrytags=industrytags)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/companies/editcompany')
def edit_company():
    title = request.args.get('title')
    if (title == '' or title is None):
        return 'no company found'
    try:
        companydocument = CompanyDocumentModel.get_row_by_title(title)
        # searchqueries = SuperSearchQueryModel.query.all()
        # report = ReportModel.query.filter_by(title='default: ' + title).first()
        industrytags = IndustryTags.query.all()
        # score = companydocument.query_score
        # if (score):
        #     score = eval(score)
        return render_template('addeditcompanies.html', companydocument=companydocument, industrytags=industrytags)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/companies/deletecompany')
def delete_company():
    title = request.args.get('title')
    if (title == '' or title is None):
        return 'no company found'
    try:
        if (CompanyDocumentModel.delete(title=title)):
            # if (ReportModel.query.filter_by(title='default: ' + title).first()):
            #     print(11111, file=sys.stderr)
            #     ReportModel.query.filter_by(title='default: ' + title).delete()
            db.session.commit()
            return redirect(url_for('companies'))
        else:
            return 'error'
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/companies/seeclassifiedsentences')
def see_classified_sentences():
    title = request.args.get('title')
    try:
        companydocument = CompanyDocumentModel.get_row_by_title(title)
        d = eval(companydocument.classified_sentences)
        return render_template('seeclassifiedsentences.html', classified_sentences=companydocument.classified_sentences)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/companies/savecompany', methods=['POST'])
def savecompany():
    try:
        result = request.form
        old_title = result.get('old_title')
        title = result.get('title')
        clean_text = result.get('clean_text')
        industry_tags = result.get('industry_tags')
        # reference_to_search_query = result.get('reference_to_search_query')
        # run_query_score = result.get('run_query_score')
        if (title is None or title == ''):
            return "title cannot be empty"
        if (clean_text is None or clean_text == ''):
            return "clean text cannot be empty"
        # clean_text = re.sub(r"(\r\n|\r|\n)","",clean_text)
        # clean_text = re.sub(r"([^0-9]\.)",r"\1 ",clean_text)

        # if(industry_tags is None or industry_tags==''):
        #     return "industry tags cannot be empty"
        # if (reference_to_search_query is None or reference_to_search_query == ''):
        #     return "reference_to_search_query cannot be empty"
        # if (reference_to_search_query == 'empty'):
        #     reference_to_search_query = None
        # if (run_query_score == 'yes'):
        #     run_query_score = True
        # else:
        #     run_query_score = False
        # temp_clean_text = re.sub(r'[\n]',' ',clean_text)
        # clean_text = re.sub(r'[^a-zA-Z0-9. ]','',temp_clean_text)

        # classified_sentences = temp_azure(clean_text)
        res = requests.post(url_for('classified', _external=True), json={"mytext": clean_text})
        if res.ok:
            classified_sentences = str(res.json())
        else:
            classified_sentences = None
        # companydocument = CompanyDocumentModel(title=title,clean_text=clean_text,classified_sentences=str(classified_sentences),industry_tags=industry_tags,reference_to_search_query=reference_to_search_query)
        companydocument = CompanyDocumentModel(title=title, clean_text=clean_text,
                                               classified_sentences=str(classified_sentences),
                                               industry_tags=industry_tags)
        if (old_title is not None and old_title != ''):
            # CompanyDocumentModel.query.filter_by(title = old_title).update(dict(title=title,clean_text=clean_text,classified_sentences=str(classified_sentences),industry_tags=industry_tags,reference_to_search_query=reference_to_search_query))
            CompanyDocumentModel.query.filter_by(title=old_title).update(
                dict(title=title, clean_text=clean_text, classified_sentences=str(classified_sentences),
                     industry_tags=industry_tags))
            db.session.commit()
        else:
            db.session.add(companydocument)
            db.session.commit()
        # report = ReportModel.query.filter_by(title='default: ' + title).first()
        # if (reference_to_search_query is not None):
        #     if (run_query_score):
        #         if (report is None):
        #             now = datetime.datetime.now(tz)
        #             one_month_ago = now - relativedelta(months=1)
        #             dt_string = now.strftime('%Y-%m-%d')
        #             one_month_ago = one_month_ago.strftime('%Y-%m-%d')
        #             default_date_from = datetime.datetime.strptime(one_month_ago, '%Y-%m-%d').date()
        #             default_date_to = datetime.datetime.strptime(dt_string, '%Y-%m-%d').date()
        #
        #             report = ReportModel(first=title, second=reference_to_search_query, frequency='Weekly',
        #                                  type='vssearchquery', status='running', title='default: ' + title,
        #                                  up_to_date=True, range_from=0, range_to=100, dimension='all', descending=True
        #                                  , date_from=default_date_from, date_to=default_date_to, total=0, current_number=0,
        #                                  running=True, date_created=datetime.datetime.now(tz))
        #             db.session.add(report)
        #             db.session.commit()
        #             executor.submit(report_background, report.id, 'vssearchquery', title, reference_to_search_query, 0,
        #                             100)
        #         else:
        #             report.second = reference_to_search_query
        #             report.status = 'running'
        #             report.running = True
        #             db.session.commit()
        #             executor.submit(report_background, report.id, 'vssearchquery', title, reference_to_search_query, 0,
        #                             100)
        # else:
        #     try:
        #         CompanyDocumentModel.query.filter_by(title=old_title).update(dict(query_score=None))
        #         if (report):
        #             ReportModel.query.filter_by(title='default: ' + title).delete()
        #             SentenceModel.delete(f_id=report.id)
        #             SentenceTextModel.delete(f_id=report.id)
        #         db.session.commit()
        #     except:
        #         print('no delete', file=sys.stderr)
        return redirect(url_for('companies'))

    except Exception as e:
        print(e, file=sys.stderr)
        return "Error"


def newdocumentadd(i, f_title, f_id):
    thread = i.get('thread')
    reach = thread.get('reach')
    social = thread.get('social')
    # entities = i.get('entities')
    site_categories = thread.get('site_categories')
    external_links = i.get('external_links')
    external_images = i.get('external_images')
    # persons = None
    # organizations = None
    # locations = None

    # if (entities):
    # persons = entities.get('persons')
    # organizations = entities.get('organizations')
    # locations = entities.get('locations')

    newdocument = NewDocumentModel(f_id=f_id, thread_uuid=thread.get('uuid'), uuid=i.get('uuid'),
                                   ord_in_thread=i.get('ord_in_thread'), parent_url=i.get('parent_url')
                                   , highlight_text=i.get('highlightText'), highlight_title=i.get('highlightTitle')
                                   , highlight_thread_title=i.get('highlightThreadTitle'), rating=i.get('rating')
                                   , crawled=i.get('crawled'), updated=i.get('updated'),
                                   site_full=thread.get('site_full')
                                   , site_section=thread.get('site_section'), section_title=thread.get('section_title')
                                   , language=i.get('language')
                                   , author=i.get('author'), text=i.get('text'), url=i.get('url'),
                                   site=thread.get('site')
                                   , title=thread.get('title'), f_title=f_title, title_full=thread.get('title_full')
                                   , published=thread.get('published'), replies_count=thread.get('replies_count')
                                   , participants_count=thread.get('participants_count'),
                                   site_type=thread.get('site_type')
                                   , country=thread.get('country'), spam_score=thread.get('spam_score')
                                   , main_image=thread.get('main_image'),
                                   performance_score=thread.get('performance_score')
                                   , domain_rank=thread.get('domain_rank'))
    if (reach):
        newdocument.reach_per_m = reach.get('per_million')
        newdocument.reach_updated = reach.get('updated')
        page_views = reach.get('page_views')
        if (page_views):
            newdocument.reach_views_per_m = page_views.get('per_million')
            newdocument.reach_views_per_u = page_views.get('per_user')

    if (social):
        facebook = social.get('facebook')
        if (facebook):
            newdocument.facebook_likes = facebook.get('likes')
            newdocument.facebook_comments = facebook.get('comments')
            newdocument.facebook_shares = facebook.get('shares')
        gplus = social.get('gplus')
        if (gplus):
            newdocument.gplus_shares = gplus.get('shares')
        pinterest = social.get('pinterest')
        if (pinterest):
            newdocument.pinterest_shares = pinterest.get('shares')
        linkedin = social.get('linkedin')
        if (linkedin):
            newdocument.linkedin_shares = linkedin.get('shares')
        stumbledupon = social.get('stumbledupon')
        if (stumbledupon):
            newdocument.stumbledupon_shares = stumbledupon.get('shares')
        vk = social.get('vk')
        if (vk):
            newdocument.vk_shares = vk.get('shares')

    json_response = get_domain_authority([newdocument.url])
    results = json_response.get('results')
    domain_authority = -1
    bucket = '-1'
    if results:
        length = len(results)
        i = 0
        while i < length:
            domain_authority = json_response['results'][i]['domain_authority']
            if domain_authority:
                if domain_authority < 10:
                    bucket = '0-10'
                elif domain_authority < 20:
                    bucket = '10-20'
                elif domain_authority < 30:
                    bucket = '20-30'
                elif domain_authority < 40:
                    bucket = '30-40'
                elif domain_authority < 50:
                    bucket = '40-50'
                elif domain_authority < 60:
                    bucket = '50-60'
                elif domain_authority < 70:
                    bucket = '60-70'
                elif domain_authority < 80:
                    bucket = '70-80'
                elif domain_authority < 90:
                    bucket = '80-90'
                elif domain_authority <= 100:
                    bucket = '90-100'
            i += 1
    newdocument.domain_authority = domain_authority
    newdocument.bucket = bucket
    db.session.add(newdocument)
    db.session.flush()
    database = []

    try:
        article = nlp(newdocument.text)
        items = [x.text for x in article.ents]
        a = Counter(items)
        for i in a:
            n = NewDocumentEntitiesModel(f_id=newdocument.id, name=i, count=a[i])
            db.session.add(n)
        db.session.commit()
    except:
        db.session.rollback()

    # if (persons):
    #     for p in persons:
    #         temp_p = NewDocumentPersonsModel(f_id=newdocument.id, name=p.get('name'), sentiment=p.get('sentiment'))
    #         database.append(temp_p)
    # if (organizations):
    #     for o in organizations:
    #         temp_o = NewDocumentOrganizationsModel(f_id=newdocument.id, name=o.get('name'),
    #                                                sentiment=o.get('sentiment'))
    #         database.append(temp_o)
    # if (locations):
    #     for l in locations:
    #         temp_l = NewDocumentLocationsModel(f_id=newdocument.id, name=l.get('name'), sentiment=l.get('sentiment'))
    #         database.append(temp_l)
    if (site_categories):
        for category in site_categories:
            temp_c = NewDocumentSiteCategoriesModel(f_id=newdocument.id, category=category)
            database.append(temp_c)
    if (external_links):
        for links in external_links:
            temp_links = NewDocumentExternalLinksModel(f_id=newdocument.id, url=links)
            database.append(temp_links)
    if (external_images):
        for image in external_images:
            temp_images = NewDocumentExternalImagesModel(f_id=newdocument.id, url=str(image))
            database.append(temp_images)

    try:
        # db.session.add(newdocument)
        for i in database:
            db.session.add(i)
    except:
        db.session.rollback()
    finally:
        db.session.commit()


@app.route('/killsearchquery')
def kill_search_query():
    id = request.args.get('id')
    if not id:
        return 'id error'
    search_query = SuperSearchQueryModel.query.filter_by(id=id).first()
    if not search_query:
        return 'search query not found error'
    search_query.running = False
    docs = SearchQueryDocumentModel.query.filter_by(f_title=search_query.title).all()
    if docs:
        search_query.current_number = len(docs)
        search_query.total = len(docs)
    else:
        search_query.current_number = 0
        search_query.total = 0
    search_query.date_completed = datetime.datetime.now()
    db.session.commit()
    return redirect(url_for('processes'))


def search_query_documents_background(id):
    f_title = SuperSearchQueryModel.query.filter_by(id=id).first().title
    search_queries = SearchQueryModel.query.filter_by(f_id=id).all()
    db.session.expunge_all()
    for searchquery in search_queries:
        try:
            today = datetime.datetime.utcnow().date()
            epoch = (today - datetime.timedelta(days=int(searchquery.freshness))).strftime('%s')
            l = ' language:' + str(searchquery.market_language_code)
            c = ' thread.country:' + str(searchquery.country_code)
            ch = ' num_chars:>' + str(searchquery.characters)
            s = ' site:' + str(searchquery.site)

            if (not searchquery.market_language_code or searchquery.market_language_code == ''):
                l = ''
            if (not searchquery.country_code or searchquery.country_code == ''):
                c = ''
            if (not searchquery.characters or searchquery.characters == ''):
                ch = ''
            if (not searchquery.site or searchquery.site == ''):
                s = ''
            query_params = {"q": "{}{}{}{}{}".format(searchquery.query_string, l, c, ch, s),
                            "ts": epoch,
                            "sort": "crawled"}
            output = webhoseio.query("filterWebContent", query_params)
            super_query = SuperSearchQueryModel.query.filter_by(id=id).first()
            if not super_query.running:
                print('killed')
                return
            if not super_query.total:
                super_query.total = 0
            super_query.total += output['totalResults']
            db.session.commit()
        except Exception as e:
            print(e, 1)
    for searchquery in search_queries:
        try:
            today = datetime.datetime.utcnow().date()
            epoch = (today - datetime.timedelta(days=int(searchquery.freshness))).strftime('%s')
            l = ' language:' + str(searchquery.market_language_code)
            # st = ' site_type:' + searchquery.site_type
            c = ' thread.country:' + str(searchquery.country_code)
            ch = ' num_chars:>' + str(searchquery.characters)
            s = ' site:' + str(searchquery.site)

            if (not searchquery.market_language_code or searchquery.market_language_code == ''):
                l = ''
            # if (not searchquery.site_type or searchquery.site_type == ''):
            #     st = ''
            if (not searchquery.country_code or searchquery.country_code == ''):
                c = ''
            if (not searchquery.characters or searchquery.characters == ''):
                ch = ''
            if (not searchquery.site or searchquery.site == ''):
                s = ''
            query_params = {"q": "{}{}{}{}{}".format(searchquery.query_string, l, c, ch, s),
                            "ts": epoch,
                            "sort": "crawled"}
            print('start')
            print(query_params)
            output = webhoseio.query("filterWebContent", query_params)
            print(output['totalResults'])
        except Exception as e:
            print(e, 1)
        while True:
            for i in output['posts']:
                try:
                    super_query = SuperSearchQueryModel.query.filter_by(id=id).first()
                    if not super_query.running:
                        print('killed')
                        return
                    super_query.current_number += 1
                    db.session.commit()
                    print(super_query.total)
                    try:
                        image = i.get('thread').get('main_image')
                    except:
                        image = 'unavailable'
                    if i.get('thread'):
                        site = i.get('thread').get('site')
                    else:
                        site = ''
                    polarity, sentiment = get_sentiment(i.get('text'))
                    temp_emotions = te.get_emotion(i.get('text'))
                    emotions = {}
                    if temp_emotions:
                        for key in temp_emotions:
                            if temp_emotions[key] != 0.0:
                                emotions[key] = temp_emotions[key]
                    searchquerydocument = SearchQueryDocumentModel(f_title=f_title, title=i.get('title'),
                                                                   author=str(i.get('author')),
                                                                   provider=str(site),
                                                                   url=i.get('url'), image_url=image,
                                                                   date=i.get('published'), clean_text=i.get('text'),
                                                                   polarity=polarity, sentiment=sentiment,
                                                                   emotions=str(emotions),
                                                                   date_created=datetime.datetime.now())
                    if (SearchQueryDocumentModel.query.filter_by(f_title=f_title,
                                                                 url=searchquerydocument.url).first() is None):
                        if len(searchquerydocument.clean_text) > 50000:
                            if super_query.running:
                                print('>50000')
                                super_query.total -= 1
                                super_query.current_number -= 1
                                db.session.commit()
                            db.session.close()
                            continue
                        res = requests.post('http://13.82.225.206:5000/predict',
                                            json={"mytext": searchquerydocument.clean_text})
                        if res.ok:
                            searchquerydocument.classified_sentences = str(res.json())
                        else:
                            searchquerydocument.classified_sentences = None
                        db.session.add(searchquerydocument)
                    else:
                        print('Already in database', file=sys.stderr)
                        if super_query.running:
                            super_query.total -= 1
                            super_query.current_number -= 1
                            db.session.commit()
                        db.session.close()
                        continue
                    db.session.flush()
                    newdocumentadd(i, f_title, searchquerydocument.id)
                    db.session.commit()
                    db.session.close()
                except Exception as e:
                    print(e, 123, 123, file=sys.stderr)
                    db.session.rollback()
                    db.session.commit()
                    super_query = SuperSearchQueryModel.query.filter_by(id=id).first()
                    if super_query.running:
                        super_query.total -= 1
                        super_query.current_number -= 1
                    db.session.commit()
                    continue
            output = webhoseio.get_next()
            if (int(output['moreResultsAvailable']) < 1):
                break
        db.session.commit()

    super_query = SuperSearchQueryModel.query.filter_by(id=id).first()
    res = export_result(
        {'export_type': 'search_query', 'where': str(super_query.id), 'filter': 'includes', 'search_parameter': '',
         'format': 'flat_json', 'flag_link': True})
    if res:
        print(res)
    else:
        print('error')
    super_query.running = False
    super_query.date_completed = datetime.datetime.now()
    docs = SearchQueryDocumentModel.query.filter_by(f_title=super_query.title).all()
    if docs:
        super_query.current_number = len(docs)
        super_query.total = len(docs)
    db.session.commit()
    db.session.close()
    print('end')


#### Report section

@app.route('/reports/')
def reports():
    try:
        reports = ReportModel.query.all()
        noun_reports = NounReportModel.query.all()
        for report in noun_reports:
            reports.append(report)
        reports.sort(key=lambda x: x.date_created, reverse=False)
        # d = eval(companydocument.classified_sentences)
        return render_template('reports.html', reports=reports)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/updatereport', methods=['POST'])
def update_report():
    try:
        result = request.form
        print(result, file=sys.stderr)
        # dimensions = {'aesthetic': 0, 'craftsmanship': 0, 'purpose': 0, 'narrative': 0}
        id = result.get('id')
        title = result.get('title')
        descending = result.get('descending')
        range_from = result.get('range_from')
        range_to = result.get('range_to')
        range_from = re.findall('\d+', range_from)
        range_to = re.findall('\d+', range_to)
        if (None in [range_from, range_to]):
            return 'error in range'
        else:
            range_from = int(range_from[0])
            range_to = int(range_to[0])
        if (range_from >= range_to or range_from > 100 or range_to > 100):
            return 'error in range'
        up_to_date = result.get('up_to_date')
        type = result.get('type')
        dimension = result.get('dimension')
        report = ReportModel.query.filter_by(id=id).first()
        if (report is None):
            return 'no report found'

        # if ('default: ' in report.title):
        #     default_company = report.title[9:]
        #     default_company = CompanyDocumentModel.query.filter_by(title=default_company).first()
        if (descending == 'true'):
            descending = True
        else:
            descending = False
        if (up_to_date == 'true'):
            up_to_date = True
        else:
            up_to_date = False
        # sentences = SentenceModel.query.filter_by(f_id=id).all()
        # if (sentences is None or sentences == []):
        #     pass
        # else:
        #     for dim in dimensions:
        #         num = 0
        #         total = 0
        #         for sentence in sentences:
        #             if (dim != sentence.dimension):
        #                 continue
        #             total += 1
        #             score = sentence.similarity
        #             if (score >= range_from and score <= range_to):
        #                 num += 1
        #         if (total > 0):
        #             dimensions[dim] = (num / total) * 100
        #         else:
        #             dimensions[dim] = 0
        #     dimensions['overall'] = (dimensions['aesthetic'] + dimensions['craftsmanship'] + dimensions['purpose'] +
        #                              dimensions['narrative']) / 4
        # print(dimensions, file=sys.stderr)
        ReportModel.query.filter_by(id=id).update(
            dict(dimension=dimension, descending=descending, range_from=range_from,
                 range_to=range_to, up_to_date=up_to_date, title=title, date_created=datetime.datetime.now(tz)))
        # ReportModel.query.filter_by(id=id).update(dict(score=str(dimensions),dimension=dimension,descending=descending,range_from=range_from,range_to=range_to,up_to_date=up_to_date,title=title))
        # if('default: ' in report.title):
        #     default_company.query_score = str(dimensions)
        db.session.commit()
        # update_score(id)
        return 'done'
        # d = eval(companydocument.classified_sentences)
        # return render_template('reports.html',reports=reports)
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/deletereport')
def delete_report():
    try:
        id = request.args.get('id')
        if (id is None or id == ''):
            return 'error'
        if (ReportModel.delete(id=id)):
            SentenceModel.delete(f_id=id)
            SentenceTextModel.delete(f_id=id)
            return redirect(url_for('reports'))
        else:
            return 'error'
    except Exception as e:
        print(e, file=sys.stderr)
        return 'error'


@app.route('/newreport', methods=['GET', 'POST'])
def new_report():
    if (request.method == 'GET'):
        try:
            type = request.args.get('type')
            companies = CompanyDocumentModel.query.all()
            second = CompanyDocumentModel.query.all()
            second_label = 'Company B'
            if (type == None):
                return render_template('newreport.html', first=companies, type='vscompany', second=companies,
                                       second_label=second_label)
            elif (type == 'vscompany'):
                pass
            elif (type == 'vssearchquery'):
                second = SuperSearchQueryModel.query.all()
                second_label = 'Search Query'
            elif (type == 'vstag'):
                second = IndustryTags.query.all()
                second_label = 'Industry Tag'
            else:
                return 'error'
            return render_template('newreport.html', first=companies, type=type, second=second,
                                   second_label=second_label)
        except Exception as e:
            print(e, file=sys.stderr)
            return 'error'
    else:
        try:

            result = request.form
            title = result.get('title')
            if (title is None or title == '' or title == 'None'):
                return 'title cannot be empty'
            first = result.get('first')
            second = result.get('second')
            frequency = result.get('frequency')
            type = result.get('type')
            up_to_date = result.get('up_to_date')
            date_from = result.get('date_from')
            date_to = result.get('date_to')
            try:
                date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
                date_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            except Exception as e:
                print(e)
                return 'date error'

            if (up_to_date is None or up_to_date is False):
                up_to_date = False
            else:
                up_to_date = True

            report = ReportModel(first=first, second=second, frequency=frequency, type=type, status='running',
                                 title=title, up_to_date=up_to_date, range_from=0, range_to=100, dimension='all',
                                 descending=True, date_from=date_from, date_to=date_to, total=0, current_number=0,
                                 running=True, date_created=datetime.datetime.now(tz))
            db.session.add(report)
            db.session.commit()
            executor.submit(report_background, report.id, type, first, second, 0, 100)
            return redirect(url_for('reports'))

        except Exception as e:
            print(e, file=sys.stderr)
            return 'error'


def get_providers(id, range_from, range_to):
    try:
        providers = []
        temp_sen = SentenceModel.query.filter(SentenceModel.f_id == id, SentenceModel.similarity >= range_from,
                                              SentenceModel.similarity <= range_to).all()
        p = {}
        # a = {}
        for i in temp_sen:
            if (p.get(i.provider) is not None):
                p[i.provider] += 1
            else:
                p[i.provider] = 1
            # if(a.get(i.author) is not None):
            #     a[i.author] += 1
            # else:
            #     a[i.author] = 1
        length_sen = len(temp_sen)
        for key, value in p.items():
            providers.append([key, (value / length_sen) * 100])
        providers = sorted(providers, key=lambda x: x[1], reverse=True)
        # for key,value in a.items():
        #     authors.append([key,value/length_sen])
        # authors = sorted(authors, key=lambda x: x[1],reverse=True)
        return providers[:5]
    except:
        return 'error'


@app.route('/loadmore', methods=['POST'])
def load_more():
    try:
        id = request.form.get('id')
        if (id is None):
            return 'error'
        offset = request.form.get('offset')
        if (offset is None or offset == ''):
            offset = 0
        else:
            offset = int(offset)
        report = ReportModel.query.filter_by(id=id).first()
        if (report is None):
            return 'report not found'
        all_sentences = SentenceModel.query.filter_by(f_id=id).all()
        sentences = []
        if (report.dimension == 'all'):
            for i in all_sentences:
                if i.similarity >= report.range_from and i.similarity <= report.range_to:
                    sentences.append(i)
        else:
            for i in all_sentences:
                if (
                        i.dimension == report.dimension and i.similarity >= report.range_from and i.similarity <= report.range_to):
                    sentences.append(i)
        if (len(sentences) > 0):
            sentences = sorted(sentences, key=lambda x: x.similarity, reverse=report.descending)
        d = []
        s = {}
        for i in SentenceTextModel.query.filter_by(f_id=report.id).all():
            s[i.id] = i.sentence
        for i in sentences[offset:]:
            try:
                print(i.sentence1, file=sys.stderr)
                d.append(
                    {'sentence1': s.get(int(i.sentence1)), 'similarity': i.similarity,
                     'sentence2': s.get(int(i.sentence2)),
                     'title': i.title2, 'id': i.id2, 'provider': i.provider, 'url': i.url})
                if len(d) >= 20:
                    break
            except:
                pass
        return {'data': d}
    except Exception as e:
        return str(e)


@app.route('/reporttest', methods=['GET', 'POST'])
def report_company_test():
    if (request.method == 'GET'):
        # try:
        dimensions = ['aesthetic', 'craftsmanship', 'purpose', 'narrative', 'all']
        id = request.args.get('id')
        chartdimension = request.args.get('chartdimension')
        if (chartdimension is None or chartdimension == '' or chartdimension == 'None'):
            chartdimension = 'all'
        if (id is None):
            return 'error'
        report = ReportModel.query.filter_by(id=id).first()
        if (report is None):
            return 'report not found'
        companydocuments = CompanyDocumentModel.query.all()
        searchqueries = SuperSearchQueryModel.query.all()
        tags = IndustryTags.query.all()
        temp_colors = ClassColors.query.filter_by(id=1).first()
        if temp_colors:
            colors = {'overall': temp_colors.overall, 'purpose': temp_colors.purpose,
                      'craftsmanship': temp_colors.craftsmanship,
                      'aesthetic': temp_colors.aesthetic, 'narrative': temp_colors.narrative}
        else:
            colors = {'overall': '', 'purpose': '', 'craftsmanship': '',
                      'aesthetic': '', 'narrative': ''}
        all_sentences = SentenceModel.query.filter(SentenceModel.f_id == id).all()
        sentences = []
        for i in all_sentences:
            if (
                    i.dimension == report.dimension and i.similarity >= report.range_from and i.similarity <= report.range_to):
                sentences.append(i)
        if (len(sentences) > 0):
            sentences = sorted(sentences, key=lambda x: x.similarity, reverse=report.descending)
        type = report.type
        page_url = ''
        score1 = None
        score2 = None
        score = None
        sentences_score = None
        providers = None
        chartdata = None
        path_to_image = None
        both = []
        tag_data = []
        # authors = []
        if (type == 'vscompany'):
            page_url = 'reportcompanytest.html'
            # first_default = ReportModel.query.filter_by(title='default: ' + report.first).first()
            # second_default = ReportModel.query.filter_by(title='default: ' + report.second).first()
            providers = []
            # if (first_default):
            #     c1 = CompanyDocumentModel.query.filter_by(title=report.first).first()
            #     if (c1 and c1.query_score):
            #         score1 = eval(c1.query_score)
            #     result1 = get_providers(id=first_default.id, range_from=first_default.range_from,
            #                             range_to=first_default.range_to)
            #     if (result1 == 'error'):
            #         providers.append([])
            #     else:
            #         providers.append(result1)
            # if (second_default):
            #     c2 = CompanyDocumentModel.query.filter_by(title=report.second).first()
            #     if (c2 and c2.query_score):
            #         score2 = eval(c2.query_score)
            #     result2 = get_providers(id=second_default.id, range_from=second_default.range_from,
            #                             range_to=second_default.range_to)
            #     if (result2 == 'error'):
            #         providers.append([])
            #     else:
            #         providers.append(result2)
            # print(providers)
        elif (type == 'vssearchquery'):
            page_url = 'reportsearchquerytest.html'
            providers = []
            if (report.score):
                # c1 = CompanyDocumentModel.query.filter_by(title=report.first).first()
                # if (c1 and c1.query_score):
                #     score1 = eval(c1.query_score)
                score = eval(report.score)
            # temp_sen = SentenceModel.query.filter(SentenceModel.f_id==report.id,SentenceModel.similarity>=report.range_from,SentenceModel.similarity<=report.range_to).all()
            # temp_sen = SentenceModel.query.filter(SentenceModel.f_id==report.id).all()
            temp_sen = []
            p = {}
            # a = {}
            print(3)

            for i in all_sentences:
                if i.similarity >= report.range_from and i.similarity <= report.range_to:
                    temp_sen.append(i)

            for i in temp_sen:
                if (p.get(i.provider) is not None):
                    p[i.provider] += 1
                else:
                    p[i.provider] = 1
                # if(a.get(i.author) is not None):
                #     a[i.author] += 1
                # else:
                #     a[i.author] = 1
            length_sen = len(temp_sen)
            for key, value in p.items():
                providers.append([key, (value / length_sen) * 100])
            providers = sorted(providers, key=lambda x: x[1], reverse=True)
            providers = providers[:5]
            print(4)
            sentences_score = get_search_query_sentence_percentage(report.second)

            print(sentences_score)
            matplotlib.use('Agg')
            matplotlib.rcParams.update({'font.size': 16})
            labels = 'Aesthetic', 'Craftsmanship', 'Narrative', 'Purpose'
            temp_colors = ClassColors.query.filter_by(id=1).first()
            if temp_colors:
                pie_colors = [temp_colors.aesthetic, temp_colors.craftsmanship, temp_colors.narrative,
                              temp_colors.purpose]
            if sentences_score:
                sizes = [sentences_score.get('aesthetic'), sentences_score.get('craftsmanship'),
                         sentences_score.get('narrative'), sentences_score.get('purpose')]
            else:
                sizes = [0, 0, 0, 0]

            fig1, ax1 = plt.subplots()
            plt.style.use('dark_background')

            if temp_colors:
                ax1.pie(sizes, autopct='%1.1f%%',
                        startangle=90, textprops={'color': "black"}, colors=pie_colors)
            else:
                ax1.pie(sizes, autopct='%1.1f%%',
                        startangle=90, textprops={'color': "black"})
            ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
            path_to_image = f"./static/images/plots/new_plot{random.randint(0, 999)}.png"
            while os.path.exists(path_to_image):
                path_to_image = f"./static/images/plots/new_plot{random.randint(0, 999)}.png"
            plt.savefig(path_to_image, transparent=True)
            # for key,value in a.items():
            #     authors.append([key,value/length_sen])
            # authors = sorted(authors, key=lambda x: x[1],reverse=True)
            print(providers, file=sys.stderr)
            # print(authors,file=sys.stderr)
        elif (type == 'vstag'):
            page_url = 'reporttagtest.html'
            # # page_url = 'test.html'
            # companies_tagged = CompanyDocumentModel.query.all()
            # documents_tagged = ArbitraryDocumentModel.query.all()
            # both_temp = CompanyDocumentModel.query.filter_by(title=report.first).first()
            # if (both_temp and companies_tagged):
            #     if (both_temp in companies_tagged):
            #         companies_tagged.remove(both_temp)
            #     companies_tagged.insert(0, both_temp)
            # else:
            #     return 'error'
            # if (companies_tagged):
            #     for i in companies_tagged:
            #         if (i.industry_tags and report.second in i.industry_tags) or (i.title == report.first):
            #             if (i.query_score):
            #                 dict_query_score = eval(i.query_score)
            #                 if dict_query_score:
            #                     for j in dict_query_score:
            #                         dict_query_score[j] = round(dict_query_score[j]['score'], 2)
            #                 tag_data.append({'query_score': dict_query_score, 'title': i.title})
            #                 both.append(i)
        else:
            'error'
        s = {}
        for i in SentenceTextModel.query.filter_by(f_id=report.id).all():
            s[i.id] = i.sentence
        result_sentences = []
        for i in sentences:
            try:
                i.sentence1 = s.get(int(i.sentence1))
                i.sentence2 = s.get(int(i.sentence2))
                result_sentences.append(i)
                if len(result_sentences) >= 20:
                    break
            except:
                pass
        try:
            date_from = int(time.mktime(report.date_from.timetuple())) * 1000
            date_to = int(time.mktime(report.date_to.timetuple())) * 1000
        except:
            date_from = None
            date_to = None
        return render_template(page_url, companydocuments=companydocuments, report=report, dimensions=dimensions,
                               sentences=result_sentences, searchqueries=searchqueries, tags=tags, score1=score1,
                               score2=score2, providers=providers, tagdata=tag_data, chartdimension=chartdimension,
                               date_from=date_from, date_to=date_to, color=colors, score=score,
                               sentences_score=sentences_score, path_to_image=path_to_image)

    # except Exception as e:
    #     print(e,file=sys.stderr)
    #     return 'error'
    else:
        try:
            dimensions = ['aesthetic', 'craftsmanship', 'purpose', 'narrative', 'all']
            result = request.form
            id = result.get('id')
            first = result.get('first')
            second = result.get('second')
            frequency = result.get('frequency')
            type = result.get('type')
            up_to_date = result.get('up_to_date')
            if (up_to_date is None or up_to_date is False):
                up_to_date = False
            else:
                up_to_date = True
            dimension = result.get('dimensions')
            # if (dimension is None):
            #     if (first is None):
            #         return 'error'
            #     report = ReportModel(first=first, second=second, frequency=frequency, type=type, status='incomplete',
            #                          up_to_date=up_to_date, total=0, current_number=0)
            #     companydocuments = CompanyDocumentModel.query.all()
            #     searchqueries = SuperSearchQueryModel.query.all()
            #     if (type == 'vscompany'):
            #         return render_template('reportcompanytest.html', companydocuments=companydocuments, report=report,
            #                                new=True, dimensions=dimensions)
            #     elif (type == 'vssearchquery'):
            #         return render_template('reportsearchquerytest.html', companydocuments=companydocuments,
            #                                searchqueries=searchqueries, report=report, new=True, dimensions=dimensions)
            #     elif (type == 'vstag'):
            #         pass
            #     else:
            #         return 'error'
            title = result.get('title')
            descending = result.get('descending')
            range_from = result.get('range_from')
            range_to = result.get('range_to')
            range_from = re.findall('\d+', range_from)
            range_to = re.findall('\d+', range_to)
            date_from = result.get('date_from')
            date_to = result.get('date_to')
            if type == 'vssearchquery':
                try:
                    date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
                    date_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
                except Exception as e:
                    print(e)
                    return 'date error'

            if (title is None or title == ''):
                return 'error in title'
            if (None in [range_from, range_to]):
                return 'error in range'
            else:
                range_from = int(range_from[0])
                range_to = int(range_to[0])
            if (range_from >= range_to or range_from > 100 or range_to > 100):
                return 'error in range'
            if (descending is None or descending is False):
                descending = False
            else:
                descending = True
            if (id is None or id == 'None' or id == ''):
                print(123, file=sys.stderr)
                now = datetime.datetime.now(tz)
                one_month_ago = now - relativedelta(months=1)
                dt_string = now.strftime('%Y-%m-%d')
                one_month_ago = one_month_ago.strftime('%Y-%m-%d')
                default_date_from = datetime.datetime.strptime(one_month_ago, '%Y-%m-%d').date()
                default_date_to = datetime.datetime.strptime(dt_string, '%Y-%m-%d').date()
                report = ReportModel(first=first, second=second, frequency=frequency, type=type, status='running',
                                     dimension=dimension
                                     , descending=descending, range_from=range_from, range_to=range_to, title=title,
                                     up_to_date=up_to_date, date_from=default_date_from, date_to=default_date_to,
                                     total=0, current_number=0, running=True, date_created=datetime.datetime.now(tz))
                db.session.add(report)
                db.session.commit()
                executor.submit(report_background, report.id, type, first, second, range_from, range_to)
                return redirect(url_for('reports'))

            if (ReportModel.query.filter_by(id=id).first() is not None):
                ReportModel.query.filter_by(id=id).update(
                    dict(first=first, second=second, frequency=frequency, type=type, status='running',
                         dimension=dimension
                         , descending=descending, range_from=range_from, range_to=range_to, title=title,
                         up_to_date=up_to_date, date_from=date_from, date_to=date_to, total=0, current_number=0,
                         running=True, date_created=datetime.datetime.now(tz)))
                db.session.commit()
                executor.submit(report_background, id, type, first, second, range_from, range_to)
                return redirect(url_for('reports'))
            else:
                return 'error sqlalchemy'

        except Exception as e:
            print(traceback.format_exc())
            return 'error'


def report_work_score(type, first, second, report, dimension, all_sentence2s, all_sen_pro_authors):
    sentence2 = []
    sen_pro_author = {}
    if (type == 'vscompany'):
        print("entered vscompany condition")
        second_company = CompanyDocumentModel.query.filter_by(title=second).first()
        print("second company: ", second_company)
        dict_company = []
        if (second_company.classified_sentences):
            dict_company = eval(second_company.classified_sentences)
            # print("dict_company: ", dict_company)
        for i in dict_company:
            if (dict_company[i] == dimension):
                ##sentence2.append(i)
                if (len(re.findall(r'\w+', i)) > 3):
                    sentence2.append([i, second_company.id, 'company', second_company.title, None])
            if (len(re.findall(r'\w+', i)) > 3):
                all_sentence2s.append([i, second_company.id, 'company', second_company.title, None])
    elif (type == 'vssearchquery'):
        print('entered search query')
        temp_date = SearchQueryDocumentModel.query.filter_by(f_title=second).all()
        second_searchquery = []
        for query in temp_date:
            try:
                published = datetime.datetime.strptime(query.date.split('T')[0], '%Y-%m-%d')
            except:
                continue
            if published:
                if report.date_from <= published <= report.date_to:
                    second_searchquery.append(query)
                # else:
                # print(query.title, 'removed')
        for querydocument in second_searchquery:
            dict_query = []
            if (querydocument.classified_sentences):
                dict_query = eval(querydocument.classified_sentences)
            for i in dict_query:
                if (dict_query[i] == dimension):
                    ##sentence2.append(i)
                    ##sen_pro_author[i] = {'provider':querydocument.provider,'author':querydocument.author}
                    if len(re.findall(r'\w+', i)) > 3:
                        sentence2.append([i, querydocument.id, 'searchquery', querydocument.title, querydocument.url])
                        sen_pro_author[i] = {'provider': querydocument.provider, 'author': querydocument.author}
                if len(re.findall(r'\w+', i)) > 3:
                    all_sentence2s.append([i, querydocument.id, 'searchquery', querydocument.title, querydocument.url])
                    all_sen_pro_authors[i] = {'provider': querydocument.provider,
                                              'author': querydocument.author}
    elif (type == 'vstag'):
        companies_tagged = CompanyDocumentModel.query.all()
        documents_tagged = ArbitraryDocumentModel.query.all()
        both = []
        if (documents_tagged):
            for i in documents_tagged:
                if (i.industry_tags and second in i.industry_tags):
                    both.append(i)
        if (companies_tagged):
            for i in companies_tagged:
                if (i.industry_tags and second in i.industry_tags):
                    both.append(i)
        if (first in both):
            both.remove(first)
        for querydocument in both:
            d = []
            if (querydocument.classified_sentences):
                d = eval(querydocument.classified_sentences)
            for i in d:
                if (d[i] == dimension):
                    ##sentence2.append(i)
                    if (len(re.findall(r'\w+', i)) > 3):
                        sentence2.append([i, querydocument.id, 'tag', querydocument.title, None])
                if (len(re.findall(r'\w+', i)) > 3):
                    all_sentence2s.append([i, querydocument.id, 'tag', querydocument.title, None])
    return sentence2, sen_pro_author


@app.route('/killreport')
def kill_report():
    id = request.args.get('id')
    if not id:
        return 'id error'
    report = ReportModel.query.filter_by(id=id).first()
    if not report:
        return 'report not found error'
    report.running = False
    report.status = 'killed'
    report.date_completed = datetime.datetime.now()
    db.session.commit()
    return redirect(url_for('processes'))


def report_background(id, type, first, second, range_from, range_to):
    try:
        print(id, type, first, second, range_from, range_to, file=sys.stderr)
        dimensions = {'aesthetic': {'num': 0, 'total': 0, 'score': 0},
                      'craftsmanship': {'num': 0, 'total': 0, 'score': 0},
                      'purpose': {'num': 0, 'total': 0, 'score': 0},
                      'all': {'num': 0, 'total': 0, 'score': 0},
                      'narrative': {'num': 0, 'total': 0, 'score': 0}}
        report = ReportModel.query.filter_by(id=id).first()
        # default_flag = False
        # if 'default: ' in report.title:
        #     default_flag = True
        first = CompanyDocumentModel.query.filter_by(title=first).first()
        dict_company_A = eval(first.classified_sentences)
        if (SentenceModel.query.filter_by(f_id=id).first() is not None):
            if (SentenceModel.delete(f_id=id) and SentenceTextModel.delete(f_id=id)):
                pass
            else:
                print('delete error', file=sys.stderr)

        process_s1_total = []
        process_s2_total = []
        process_all_sen_pro_authors = {}
        process_total = 0
        for dimension in dimensions:
            process_s1 = []
            for i in dict_company_A:
                if (dict_company_A[i] == dimension):
                    ##sentence1.append(i)
                    if (len(re.findall(r'\w+', i)) > 3):
                        process_s1.append(i)
                if (len(re.findall(r'\w+', i)) > 3):
                    process_s1_total.append(i)
            process_s2, sen_pro_author = report_work_score(type, first, second, report, dimension, process_s2_total,
                                                           process_all_sen_pro_authors)
            s2 = []
            for i in process_s2:
                s2.append(i[0])
            process_total += (len(list(dict.fromkeys(process_s1))) * len(list(dict.fromkeys(s2))))
        s2 = []
        for i in process_s2_total:
            s2.append(i[0])
        process_total += (len(list(dict.fromkeys(process_s1_total))) * len(list(dict.fromkeys(s2))))
        print(process_total)
        report.total = process_total
        db.session.commit()

        all_sentence1s = []
        all_sentence2s = []
        all_sen_pro_authors = {}
        for dimension in dimensions:

            if not ReportModel.query.filter_by(id=id).first().running:
                print('killed')
                return
            sentence1 = []
            for i in dict_company_A:
                if (dict_company_A[i] == dimension):
                    ##sentence1.append(i)
                    if (len(re.findall(r'\w+', i)) > 3):
                        sentence1.append(i)
                if (len(re.findall(r'\w+', i)) > 3):
                    all_sentence1s.append(i)
            sentence2, sen_pro_author = report_work_score(type, first, second, report, dimension, all_sentence2s,
                                                          all_sen_pro_authors)
            if (len(sentence1) == 0 or len(sentence2) == 0):
                continue
            # print(dimension,sentence1,sentence2,file=sys.stderr)
            temp_score = get_scores(sentence1, sentence2, dimension, id, sen_pro_author)
            # if temp_score and default_flag:
            if temp_score:
                num = temp_score.get('num')
                total = temp_score.get('total')
                try:
                    dimensions[dimension] = {'num': num, 'total': total, 'score': (num / total) * 100}
                except:
                    pass

        # scoring for comparing all sentences
        if len(all_sentence1s) > 0 and len(all_sentence2s) > 0:
            if not ReportModel.query.filter_by(id=id).first().running:
                print('killed')
                return
            dimension = 'all'
            temp_all_score = get_scores(all_sentence1s, all_sentence2s, dimension, id,
                                        sen_pro_author=all_sen_pro_authors)
            # if temp_all_score and default_flag:
            if temp_all_score:
                num = temp_all_score.get('num')
                total = temp_all_score.get('total')
                try:
                    dimensions[dimension] = {'num': num, 'total': total, 'score': (num / total) * 100}
                except:
                    pass
                ReportModel.query.filter_by(id=id).update(dict(score=str(dimensions)))
                # first.query_score = str(dimensions)
        try:
            d_f = report.date_from.strftime('%Y-%m-%d')
            d_t = report.date_to.strftime('%Y-%m-%d')
        except:
            d_f = None
            d_t = None
        if d_f is None and d_t is None:

            res = export_result(
                {'export_type': 'report', 'where': str(report.id), 'filter': 'includes',
                 'search_parameter': '',
                 'format': 'json', 'flag_link': True})

            # res = requests.post(url_for('export_result', _external=True),
            #                     data=[('export_type', 'report'), ('where', str(report.id)),
            #                           ('filter', 'includes'), ('search_parameter', ''), ('format', 'json'),
            #                           ('flag_link', True)])
        else:

            res = export_result(
                {'export_type': 'report', 'where': str(report.id), 'date_checkbox': 'date', 'filter': 'includes',
                 'search_parameter': '', 'start_date': d_f, 'end_date': d_t,
                 'format': 'json', 'flag_link': True})

            # res = requests.post(url_for('export_result', _external=True),
            #                     data=[('export_type', 'report'), ('where', str(report.id)), ('date_checkbox', 'date'),
            #                           ('start_date', d_f), ('end_date', d_t), ('filter', 'includes'),
            #                           ('search_parameter', ''), ('format', 'json'), ('flag_link', True)])

        if res:
            print(res)
        else:
            print('error')
        ReportModel.query.filter_by(id=id).update(
            dict(status='done', running=False, date_completed=datetime.datetime.now()))
        db.session.commit()

    except Exception as e:
        print(traceback.format_exc(), file=sys.stderr)
        ReportModel.query.filter_by(id=id).update(
            dict(status='error', running=False, date_completed=datetime.datetime.now()))
        db.session.commit()


def update_score(update_id):
    dimensions = {'aesthetic': {'num': 0, 'total': 0, 'score': 0},
                  'craftsmanship': {'num': 0, 'total': 0, 'score': 0},
                  'purpose': {'num': 0, 'total': 0, 'score': 0},
                  'all': {'num': 0, 'total': 0, 'score': 0},
                  'narrative': {'num': 0, 'total': 0, 'score': 0}}
    report = ReportModel.query.filter_by(id=update_id).first()
    # if 'default: ' not in report.title:
    #     return None
    first = CompanyDocumentModel.query.filter_by(title=report.first).first()
    thresh = Threshold.query.filter_by(id=1).first()
    if thresh:
        value = thresh.value
    else:
        value = 100
    db.session.commit()
    sentences = SentenceModel.query.filter_by(f_id=update_id).all()
    report_score = eval(report.score)
    for dimension in dimensions:
        report_total = report_score[dimension]['total']
        num = 0
        # total = 0
        for sentence in sentences:
            if sentence.dimension != dimension:
                continue
            # total += 1
            if sentence.similarity >= value:
                num += 1
        if report_total != 0:
            dimensions[dimension] = {'num': num, 'total': report_total, 'score': (num / report_total) * 100}

    # overall_num = dimensions['aesthetic']['num'] + dimensions['craftsmanship']['num'] + dimensions['purpose']['num'] + \
    #               dimensions['narrative']['num']
    # overall_total = dimensions['aesthetic']['total'] + dimensions['craftsmanship']['total'] + dimensions['purpose'][
    #     'total'] + dimensions['narrative']['total']
    # try:
    #     dimensions['overall'] = {
    #         'num': overall_num,
    #         'total': overall_total, 'score': (overall_num / overall_total) * 100}
    # except:
    #     dimensions['overall'] = {'num': 0, 'total': 0, 'score': 0}

    report.score = str(dimensions)
    db.session.commit()
    # if ('default: ' in ReportModel.query.filter_by(id=update_id).first().title):
    #     first.query_score = str(dimensions)
    #     db.session.commit()


def get_scores(sentence1, sentence2, dimension, id, sen_pro_author):
    print('get scores')
    threshold = Threshold.query.filter_by(id=1).first()
    if threshold:
        value = threshold.value
    else:
        value = 100
    for s in sentence1:
        f = SentenceTextModel.query.filter_by(f_id=id, sentence=s).first()
        if (f is None):
            polarity, sentiment = get_sentiment(s)
            temp_emotions = te.get_emotion(s)
            emotions = {}
            if temp_emotions:
                for key in temp_emotions:
                    if temp_emotions[key] != 0.0:
                        emotions[key] = temp_emotions[key]
            s = SentenceTextModel(f_id=id, sentence=s, polarity=polarity, sentiment=sentiment, emotions=str(emotions))
            db.session.add(s)
            db.session.commit()
    for s in sentence2:
        if (SentenceTextModel.query.filter_by(f_id=id, sentence=s[0]).first() is None):
            polarity, sentiment = get_sentiment(s[0])
            temp_emotions = te.get_emotion(s[0])
            emotions = {}
            if temp_emotions:
                for key in temp_emotions:
                    if temp_emotions[key] != 0.0:
                        emotions[key] = temp_emotions[key]
            s = SentenceTextModel(f_id=id, sentence=s[0], polarity=polarity, sentiment=sentiment,
                                  emotions=str(emotions))
            db.session.add(s)
            db.session.commit()

    res_dict = getSimlarity(sentence1, sentence2)
    t = []
    st = SentenceTextModel.query.filter_by(f_id=id).all()
    s = {}
    for i in st:
        s[i.sentence] = i.id
    # print(s)
    num = 0
    total = 0
    report_temp = ReportModel.query.filter_by(id=id).first()
    counter = report_temp.current_number
    if not counter:
        counter = 0
    for i in res_dict:
        for j in res_dict[i]:
            counter += 1
            if not s.get(j) or not s.get(i):
                continue
            total += 1
            score = abs(float(res_dict[i][j]['similarity']) * 100)
            if score >= value:
                num += 1
                if (sen_pro_author == {}):
                    db.session.add(SentenceModel(sentence1=s.get(i), sentence2=s.get(j), similarity=int(score), f_id=id,
                                                 dimension=dimension, title2=res_dict[i][j]['title'],
                                                 id2=res_dict[i][j]['id'],
                                                 type=res_dict[i][j]['type'], url=res_dict[i][j]['url']))
                else:
                    db.session.add(SentenceModel(sentence1=s.get(i), sentence2=s.get(j), similarity=int(score), f_id=id,
                                                 dimension=dimension, title2=res_dict[i][j]['title'],
                                                 id2=res_dict[i][j]['id'],
                                                 type=res_dict[i][j]['type'],
                                                 provider=sen_pro_author.get(j).get('provider'),
                                                 author=sen_pro_author.get(j).get('author'), url=res_dict[i][j]['url']))
        report_temp.current_number = counter
        db.session.commit()
        print(counter)
    # try:
    #     db.session.add_all(list(dict.fromkeys(t)))
    #     db.session.commit()
    # except Exception as e:
    #     print(e, file=sys.stderr)
    if total > 0:
        return {'num': num, 'total': total}
    else:
        return 0


def temp_azure(tmp):
    import random
    l = tmp.split('.')
    d = {}
    classes = ['purpose', 'craftsmanship', 'aesthetic', 'narrative']
    for i in l:
        if (len(i) > 5):
            d[i + '.'] = random.choice(classes)

    return d


def send_email(content):
    mail_content = content
    sender_address = 'saadcheemaa545@gmail.com'
    sender_pass = 'vtrxlybjsutiphgz'
    # receiver_address = 'hbutt877877@gmail.com'
    receiver_address = 'metalindustries8@gmail.com'
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] = 'PakMetalIndustries New Order'
    message.attach(MIMEText(mail_content, 'plain'))
    mail_session = smtplib.SMTP('smtp.gmail.com', 587)
    mail_session.starttls()
    mail_session.login(sender_address, sender_pass)
    text = message.as_string()
    mail_session.sendmail(sender_address, receiver_address, text)
    mail_session.quit()
    print('Mail Sent')


@app.route("/sendmail", methods=[POST])
def sendmail():
    try:
        result = request.get_json(force=True)
        name = result.get('name')
        address = result.get('address')
        contact = result.get('contact')
        items = result.get('items')
        print(name, items)
        text = f'You got a new order. Here are the details:\nName: {name}\nAddress: {address}\nContact: {contact}\n' \
               f'items: {items}'
        send_email(text)
        return 'sent'
    except:
        return 'error'


@app.route("/edgebundling", methods=[GET, POST])
def edge_bundling():
    report_id = request.args.get('report_id')
    report_name = f'./static/jsons/report{report_id}.json'
    if not os.path.exists(report_name):
        return 'no file exists'
    return render_template('index.html', report_name=report_name)


@app.route("/pivot", methods=[GET, POST])
def pivot():
    query_id = request.args.get('query_id')
    query_name = f'./static/jsons/searchquery{query_id}.json'
    if not os.path.exists(query_name):
        return 'no file exists'
    return render_template('pivot.html', query_name=query_name)


@app.route("/processes", methods=[GET, POST])
def processes():
    try:
        processes = []
        searchqueries = SuperSearchQueryModel.query.filter_by(running=True).all()
        reports = ReportModel.query.filter_by(running=True).all()
        noun_reports = NounReportModel.query.filter_by(running=True).all()
        for report in noun_reports:
            reports.append(report)
        recent = []
        last_hour = datetime.datetime.now() - datetime.timedelta(hours=1)
        for r in ReportModel.query.filter_by(running=False).all():
            if r.date_completed and r.date_completed > last_hour:
                recent.append(r)
        for r in NounReportModel.query.filter_by(running=False).all():
            print(r.date_completed, last_hour)
            if r.date_completed and r.date_completed > last_hour:
                print(2)
                recent.append(r)
        for s in SuperSearchQueryModel.query.filter_by(running=False).all():
            if s.date_completed and s.date_completed > last_hour:
                recent.append(s)
        for s in searchqueries:
            processes.append(s)
        for r in reports:
            processes.append(r)
        return render_template('processes.html', processes=processes, recent=recent)
    except Exception as e:
        print(e)
        return 'error'


def get_search_query_sentence_percentage(f_title):
    try:

        docs = SearchQueryDocumentModel.query.filter_by(f_title=f_title).all()
        if not docs:
            print('not docs ', f_title)
            return None
        score = {'aesthetic': 0, 'craftsmanship': 0, 'narrative': 0, 'purpose': 0}
        for d in docs:
            try:
                sentences = eval(d.classified_sentences)
            except:
                continue
            for s in sentences:
                dimension = sentences[s]
                if dimension == 'narrative':
                    score['narrative'] += 1
                elif dimension == 'purpose':
                    score['purpose'] += 1
                elif dimension == 'craftsmanship':
                    score['craftsmanship'] += 1
                elif dimension == 'aesthetic':
                    score['aesthetic'] += 1
        total = score['aesthetic'] + score['craftsmanship'] + score['narrative'] + score['purpose']
        if total == 0:
            print('total 0')
            return None
        score['aesthetic'] = round(score['aesthetic'] * 100 / total, 2)
        score['craftsmanship'] = round(score['craftsmanship'] * 100 / total, 2)
        score['narrative'] = round(score['narrative'] * 100 / total, 2)
        score['purpose'] = round(score['purpose'] * 100 / total, 2)
        return score
    except Exception as e:
        print(e)
        return None


@app.route("/viewentity", methods=[GET, POST])
def view_entity():
    id = request.args.get('id')
    top = request.args.get('top')
    if not id:
        return 'id error'
    if not top:
        top = 25
    else:
        try:
            top = int(top)
        except:
            top = 25
    noun_report = NounReportModel.query.filter_by(id=id).first()
    if not noun_report:
        return 'noun report not found'
    try:
        date_from = int(time.mktime(noun_report.date_from.timetuple())) * 1000
        date_to = int(time.mktime(noun_report.date_to.timetuple())) * 1000
    except:
        date_from = None
        date_to = None

    entities = []
    temp_entities = NounReportEntitiesModel.query.filter_by(noun_report_id=id, ignored=False, alias_id=None).all()
    for entity in temp_entities:
        entities.append([entity.name, entity.count])
    aliases = AliasModel.query.filter_by(noun_report_id=id).all()
    for alias in aliases:
        name = alias.name
        count = 0
        alias_entities = NounReportEntitiesModel.query.filter_by(noun_report_id=id, alias_id=alias.id).all()
        for i in alias_entities:
            count = count + i.count
        entities.append([name, count])

    if entities:
        entities.sort(key=lambda x: x[1], reverse=True)
    else:
        return render_template('viewentity.html', path_to_image=None, noun_report=noun_report,
                               date_from=date_from,
                               date_to=date_to, topstyle=[None, None, None])
    matplotlib.use('Agg')
    plt.rcParams.update(plt.rcParamsDefault)

    # matplotlib.rcParams.update({'font.size': 10})
    # plt.style.use('classic')
    names = []
    counts = []
    for entity in entities[:top]:
        names.append(entity[0])
        counts.append(entity[1])
    # print(names)
    # print(counts)
    if top == 25:
        fig = plt.figure(figsize=(5, 6))
    elif top == 50:
        fig = plt.figure(figsize=(6, 10))
    elif top == 100:
        fig = plt.figure(figsize=(7, 15))
    plt.barh(names, counts)
    plt.gca().invert_yaxis()
    plt.tight_layout()
    # plt.title('Store Inventory')
    # plt.ylabel('Product')
    # plt.xlabel('Quantity')
    # plt.show()
    path_to_image = f"./static/images/plots/new_plot{random.randint(0, 999)}.png"
    while os.path.exists(path_to_image):
        path_to_image = f"./static/images/plots/new_plot{random.randint(0, 999)}.png"
    plt.savefig(path_to_image, transparent=True)

    topstyle = []
    if top == 25:
        topstyle.append('font-weight: bold;')
        topstyle.append(None)
        topstyle.append(None)
    elif top == 50:
        topstyle.append(None)
        topstyle.append('font-weight: bold;')
        topstyle.append(None)
    elif top == 100:
        topstyle.append(None)
        topstyle.append(None)
        topstyle.append('font-weight: bold;')

    return render_template('viewentity.html', path_to_image=path_to_image, noun_report=noun_report, date_from=date_from,
                           date_to=date_to, topstyle=topstyle)


@app.route("/editentity", methods=[GET])
def edit_entity():
    id = request.args.get('id')
    if not id:
        return 'id error'
    noun_report = NounReportModel.query.filter_by(id=id).first()
    if not noun_report:
        return 'noun report not found'
    entities = NounReportEntitiesModel.query.filter_by(noun_report_id=id, ignored=False, alias_id=None).all()
    ignored = NounReportEntitiesModel.query.filter_by(noun_report_id=id, ignored=True).all()
    aliases = AliasModel.query.filter_by(noun_report_id=id).all()
    zipped_aliases = []
    for alias in aliases:
        alias_entities = NounReportEntitiesModel.query.filter_by(noun_report_id=id, alias_id=alias.id).all()
        zipped_aliases.append([alias, alias_entities])
    if entities:
        try:
            entities.sort(key=lambda x: x.count, reverse=True)
        except:
            pass
    # else:
    #     return 'no entities found'
    try:
        date_from = int(time.mktime(noun_report.date_from.timetuple())) * 1000
        date_to = int(time.mktime(noun_report.date_to.timetuple())) * 1000
    except:
        date_from = None
        date_to = None

    return render_template('editnounreport.html', entities=entities, noun_report=noun_report, ignored=ignored,
                           zipped_aliases=zipped_aliases, date_from=date_from, date_to=date_to)


@app.route("/savenounreport", methods=[POST])
def save_noun_report():
    data = request.get_json()
    id = data.get('id')
    entities = data.get('entities')
    ignored = data.get('ignored')
    alias = data.get('alias')
    alias_headings = data.get('alias_headings')
    if len(alias) != len(alias_headings):
        return 'error'
    NounReportEntitiesModel.query.filter_by(noun_report_id=id).delete()
    AliasModel.query.filter_by(noun_report_id=id).delete()
    db.session.flush()
    for entity in entities:
        name, count = entity.split('--->')
        try:
            count = int(count)
        except:
            count = 0
        add_entity = NounReportEntitiesModel(noun_report_id=id, name=name,
                                             count=count, ignored=False, alias_id=None)
        db.session.add(add_entity)
        db.session.flush()
    for entity in ignored:
        name, count = entity.split('--->')
        try:
            count = int(count)
        except:
            count = 0
        add_entity = NounReportEntitiesModel(noun_report_id=id, name=name,
                                             count=count, ignored=True, alias_id=None)
        db.session.add(add_entity)
        db.session.flush()

    for i in zip(alias, alias_headings):
        add_alias = AliasModel(name=i[1], noun_report_id=id)
        db.session.add(add_alias)
        db.session.flush()
        for j in i[0]:
            name, count = j.split('--->')
            try:
                count = int(count)
            except:
                count = 0
            add_entity = NounReportEntitiesModel(noun_report_id=id, name=name,
                                                 count=count, ignored=False, alias_id=add_alias.id)
            db.session.add(add_entity)
            db.session.flush()
    db.session.commit()
    return 'done'


@app.route('/newnounreport', methods=['GET', 'POST'])
def new_noun_report():
    if (request.method == 'GET'):
        try:
            search_queries = SuperSearchQueryModel.query.all()
            return render_template('newnounreport.html', search_queries=search_queries)
        except Exception as e:
            print(e, file=sys.stderr)
            return 'error'
    else:
        try:
            result = request.form
            id = result.get('id')
            title = result.get('title')
            if (title is None or title == '' or title == 'None'):
                return 'title cannot be empty'
            search_query = result.get('search_query')
            date_from = result.get('date_from')
            date_to = result.get('date_to')
            try:
                date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
                date_to = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            except Exception as e:
                print(e)
                return 'date error'
            print(search_query)

            if id:
                noun_report = NounReportModel.query.filter_by(id=id).first()
                noun_report.date_from = date_from
                noun_report.date_to = date_to
                noun_report.status = 'running'
                noun_report.running = True
                noun_report.date_created = datetime.datetime.now(tz)
                # NounReportEntitiesModel.query.filter_by(noun_report_id=id).delete()
                for entity in NounReportEntitiesModel.query.filter_by(noun_report_id=id).all():
                    entity.count = 0
            else:
                noun_report = NounReportModel(search_query_title=search_query, title=title, date_from=date_from,
                                              date_to=date_to, date_created=datetime.datetime.now(tz), type='noun',
                                              status='running', running=True)
                db.session.add(noun_report)
            db.session.flush()
            temp_date = NewDocumentModel.query.filter_by(f_title=search_query).all()
            docs = []
            for query in temp_date:
                try:
                    published = datetime.datetime.strptime(query.published.split('T')[0], '%Y-%m-%d').date()
                except:
                    continue
                # print(date_from, published, date_to)
                if published and date_from <= published <= date_to:
                    docs.append(query)
            for d in docs:
                doc_entities = NewDocumentEntitiesModel.query.filter_by(f_id=d.id).all()
                for doc_entity in doc_entities:
                    name_exists = NounReportEntitiesModel.query.filter_by(noun_report_id=noun_report.id,
                                                                          name=doc_entity.name).first()
                    if name_exists:
                        name_exists.count += doc_entity.count
                        db.session.flush()
                        continue
                    else:
                        add_entity = NounReportEntitiesModel(noun_report_id=noun_report.id, name=doc_entity.name,
                                                             count=doc_entity.count, ignored=False, alias_id=None)
                        db.session.add(add_entity)
                        db.session.flush()
            noun_report.status = 'done'
            noun_report.running = False
            noun_report.date_completed = datetime.datetime.now()
            NounReportEntitiesModel.query.filter_by(noun_report_id=id, count=0).delete()
            db.session.commit()
            return redirect(f'/editentity?id={noun_report.id}')

        except Exception as e:
            print(e, file=sys.stderr)
            return 'error'


@app.route('/deletenounreport')
def delete_noun_report():
    try:
        id = request.args.get('id')
        if (id is None or id == ''):
            return 'error'
        NounReportModel.query.filter_by(id=id).delete()
        NounReportEntitiesModel.query.filter_by(noun_report_id=id).delete()
        AliasModel.query.filter_by(noun_report_id=id).delete()
        db.session.commit()
        return redirect(url_for('reports'))
    except Exception as e:
        print(e, file=sys.stderr)
        return 'delete error'


@app.route('/fuzzy', methods=['GET'])
def fuzzy():
    id = request.args.get('id')
    if (id is None or id == ''):
        return 'error'
    ignore_list = ['one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten', 'eleven', 'twelve',
                   'thirteen', 'fourteen', 'fifteen', 'sixteen', 'seventeen', 'eighteen', 'nineteen', 'twenty',
                   'thirty', 'forty', 'fifty', 'hundred', 'thousand', 'million', 'billion', 'trillion', 'winter',
                   'january', 'february', 'march', 'spring', 'april', 'may', 'june',
                   'summer', 'july', 'august', 'september', 'autumn', 'october',
                   'november', 'december', 'winter', 'second', 'minute', 'hour', 'day', 'week', 'month',
                   'quarter', 'half', 'year', 'decade', 'century', 'millenia', 'period', 'season', 'dollar', 'usd',
                   'pound', 'euro', 'yuan', 'rupee', 'pkr', 'sterling', '%', 'percent', 'degree', 'celsius',
                   'fahrenheit', 'tomorrow', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday',
                   'sunday']
    AliasModel.query.filter_by(noun_report_id=id).delete()
    entities = NounReportEntitiesModel.query.filter_by(noun_report_id=id).all()
    for entity in entities:
        if not entity.name:
            continue
        entity.ignored = False
        entity.alias_id = None
        is_ignored = len(re.findall('\d', entity.name)) > 0
        if is_ignored or 'http' in entity.name:
            entity.ignored = True
            continue
        name = ps.stem(entity.name)
        if not name:
            continue
        for w in ignore_list:
            if find_word(w)(name):
                print(w, name)
                entity.ignored = True
                break
    db.session.flush()
    entities = NounReportEntitiesModel.query.filter_by(noun_report_id=id, ignored=False).all()
    for entity in entities:
        alias_entities = []
        for entity_2 in entities:
            if not entity == entity_2 and entity_2.alias_id is None and entity.alias_id is None and not entity_2.ignored:
                ratio = fuzz.token_set_ratio(entity.name, entity_2.name)
                if ratio > 80:
                    if not alias_entities:
                        alias_entities.append(entity)
                    alias_entities.append(entity_2)
        if alias_entities:
            a = AliasModel(name=alias_entities[0].name, noun_report_id=id)
            db.session.add(a)
            db.session.flush()
            for i in alias_entities:
                i.alias_id = a.id
        db.session.flush()
    db.session.commit()
    return redirect(f'/editentity?id={id}')


def find_word(w):
    return re.compile(r'\b({0})\b'.format(w), flags=re.IGNORECASE).search


def get_domain_authority(l):
    auth = ('mozscape-5084ac4783', 'b6f939908792253946d015256924696c')
    url = 'https://lsapi.seomoz.com/v2/url_metrics'
    l_s = "["
    for s in l:
        l_s += f'"{s}",'
    l_s = l_s[:-1]
    l_s += "]"
    data = f"""
              {{"targets": {l_s}}}
            """
    response = requests.post(url, data=data, auth=auth)
    json_response = response.json()
    return json_response


if __name__ == '__main__':
    from waitress import serve

    tl.start()
    serve(app, host='0.0.0.0', port=5000)
    # app.run(
    # host = '0.0.0.0',
    # port = 5000,debug=False
    ##ssl_context = ('cert.pem', 'key.pem')
    # )
